import streamlit as st
import pandas as pd
import requests
import re
import random
import calendar
from datetime import datetime, date, timedelta
from collections import defaultdict
<<<<<<< HEAD
import io
=======
from io import BytesIO, StringIO
>>>>>>> f585c790da4d1a3b11f6ab84c1283f44dd5e6219
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
<<<<<<< HEAD

=======
>>>>>>> f585c790da4d1a3b11f6ab84c1283f44dd5e6219

st.set_page_config(page_title="Puskesmas Sangkali", page_icon="assets/logo.png", layout="wide")

st.markdown("""
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.0/css/all.min.css">
<style>
<<<<<<< HEAD
  .main-header {
    background: linear-gradient(135deg, #14532d 0%, #166534 100%);
    padding: 2rem; border-radius: 20px; color: white;
    text-align: center; margin-bottom: 2rem;
  }
  .main-header h1 { margin: 0; }
  .data-card { background: white; padding: 1.5rem; border-radius: 16px;
               border: 1px solid #e5e7eb; margin-bottom: 1rem; }
  .result-box { background: #f8fafc; padding: 1rem; border-radius: 12px;
                border-left: 5px solid #14532d; margin: 0.5rem 0; }
  .icon-text i { margin-right: 8px; color: #14532d; }
  .logo-container { display: flex; justify-content: center; margin-bottom: 1rem; }
  .badge-online { display: inline-block; background: #dcfce7; color: #15803d;
                  font-size: 0.75rem; font-weight: 600; padding: 2px 10px; border-radius: 999px; }
  .badge-offline { display: inline-block; background: #fee2e2; color: #dc2626;
                   font-size: 0.75rem; font-weight: 600; padding: 2px 10px; border-radius: 999px; }
  .svg-icon { display: inline-block; vertical-align: middle; margin-right: 6px; }
  .past { color: #9ca3af; }
  .today { color: #15803d; font-weight: bold; }
  .tomorrow { color: #2563eb; }
  .future { color: #111827; }
  .reduced-link { color: #6b7280; font-style: italic; font-size: 0.9rem; }
  .notification-card { background: #dcfce7; border: 1px solid #bbf7d0;
                       padding: 1rem; border-radius: 12px; margin-bottom: 1rem; }
  .notification-card ul { margin-top: 0.5rem; padding-left: 1.5rem; }
  .notification-card li { margin-bottom: 0.25rem; }
  .center-red-btn { display: flex; justify-content: center; margin: 1rem 0; }
  .center-red-btn button { background-color: #dc2626 !important; color: white !important;
                           border: none !important; font-weight: 500;
                           padding: 0.4rem 1.5rem !important; border-radius: 8px; width: auto !important; }
  .center-red-btn button:hover { background-color: #b91c1c !important; }
  .randomize-card { background: #f0fdf4; border: 1px solid #bbf7d0; padding: 1rem; border-radius: 12px; margin-bottom: 1rem; }
=======
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #14532d 0%, #166534 60%, #15803d 100%);
}
[data-testid="stSidebar"] * { color: white !important; }
[data-testid="stSidebar"] .stButton > button {
    background: rgba(255,255,255,0.12);
    border: 1px solid rgba(255,255,255,0.25);
    border-radius: 10px;
    font-weight: 500;
    width: 100%;
    margin-bottom: 4px;
    padding: 0.5rem 0.75rem;
}
[data-testid="stSidebar"] .stButton > button:hover { background: rgba(255,255,255,0.28); }
[data-testid="stSidebar"] [data-testid="stImage"] {
    display: flex !important;
    justify-content: center !important;
}
[data-testid="stSidebar"] [data-testid="stImage"] img { margin: 0 auto !important; }
.sidebar-puskesmas-name { font-size:0.95rem; font-weight:700; text-align:center; margin:0.4rem 0 0.25rem; width:100%; }
.sidebar-puskesmas-sub { font-size:0.75rem; text-align:center; opacity:0.75; margin-bottom:0.5rem; width:100%; }
.sidebar-menu-label { font-size:0.7rem; text-transform:uppercase; letter-spacing:0.1em; font-weight:600; padding:0.5rem 0 0.2rem; opacity:0.55; text-align:center !important; }
.sidebar-user-badge { font-size:0.82rem; text-align:center; margin:4px 0 8px; display:block; padding:4px 0; }
.badge-online { display:inline-flex; align-items:center; gap:5px; background:rgba(187,247,208,0.2); color:#bbf7d0; font-size:0.78rem; font-weight:600; padding:3px 10px; border-radius:999px; border:1px solid rgba(187,247,208,0.3); }
.badge-offline { display:inline-flex; align-items:center; gap:5px; background:rgba(254,202,202,0.2); color:#fecaca; font-size:0.78rem; font-weight:600; padding:3px 10px; border-radius:999px; border:1px solid rgba(254,202,202,0.3); }
.main-header { background:linear-gradient(135deg,rgba(20,83,45,0.85) 0%,rgba(22,101,52,0.85) 100%); padding:4rem 2rem; border-radius:20px; color:white; text-align:center; margin-bottom:2rem; }
.main-header h1 { font-size:2.5rem; font-weight:700; margin:0; text-shadow:2px 2px 4px rgba(0,0,0,0.3); }
.main-header p { font-size:1.2rem; margin-top:0.5rem; text-shadow:1px 1px 3px rgba(0,0,0,0.3); }
.receipt-box { background:#fffef8; border:1px solid #e2e0d8; border-radius:4px; padding:0; margin:1rem 0; box-shadow:0 1px 3px rgba(0,0,0,0.08),0 4px 12px rgba(0,0,0,0.06); font-family:'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; position:relative; }
.receipt-box::before { content:""; display:block; height:6px; background:radial-gradient(circle at 6px 6px,transparent 6px,#fffef8 6px) -6px 0,linear-gradient(#fffef8,#fffef8); background-size:12px 6px,100% 100%; margin-bottom:2px; }
.receipt-box::after { content:""; display:block; height:6px; background:radial-gradient(circle at 6px 0px,transparent 6px,#fffef8 6px) -6px 0; background-size:12px 6px; margin-top:2px; }
.receipt-header { background:#14532d; color:white; text-align:center; padding:1rem 1.25rem 0.85rem; }
.receipt-header h3 { margin:0; font-size:1rem; font-weight:700; text-transform:uppercase; }
.receipt-body { padding:0.5rem 1.25rem; }
.receipt-item { padding:0.75rem 0; border-bottom:1px dashed #ddd8cc; }
.receipt-item:last-child { border-bottom:none; }
.receipt-item-title { font-weight:700; color:#14532d; font-size:0.92rem; }
.receipt-item-row { display:flex; justify-content:flex-start; align-items:baseline; font-size:0.82rem; color:#4b5563; margin-bottom:0.2rem; }
.receipt-item-label { color:#6b7280; min-width:80px; margin-right:6px; }
.receipt-item-value { color:#1f2937; text-align:left; word-break:break-word; }
.receipt-penyerta-list { margin-top:0.2rem; padding-left:0; }
.receipt-penyerta-list span { display:block; margin-bottom:0.15rem; }
.receipt-status-badge { font-size:0.72rem; font-weight:700; padding:2px 9px; border-radius:3px; text-transform:uppercase; }
.status-past { background:#f3f4f6; color:#6b7280; border:1px solid #d1d5db; }
.status-today { background:#dcfce7; color:#15803d; border:1px solid #86efac; }
.status-soon { background:#dbeafe; color:#1d4ed8; border:1px solid #93c5fd; }
.status-future { background:#fef9c3; color:#854d0e; border:1px solid #fde047; }
.receipt-footer { text-align:center; padding:0.75rem 1.25rem 0.9rem; color:#9ca3af; font-size:0.75rem; border-top:2px dashed #c8c5b8; }
.receipt-barcode { font-size:1.6rem; letter-spacing:0.05em; color:#374151; margin:0.3rem 0; }
.data-card { background:white; padding:1.25rem; border-radius:14px; border:1px solid #e5e7eb; margin-bottom:1rem; box-shadow:0 2px 8px rgba(0,0,0,0.04); }
.stat-card { background:linear-gradient(135deg,#f0fdf4,#dcfce7); border:1px solid #bbf7d0; padding:1.1rem 1.25rem; border-radius:14px; text-align:center; }
.stat-card-num { font-size:1.8rem; font-weight:700; color:#15803d; }
.history-item { background:white; border:1px solid #e5e7eb; border-radius:12px; padding:0.85rem 1rem; margin-bottom:0.6rem; }
.history-item-hadir { border-left:5px solid #16a34a; }
.history-item-tidak { border-left:5px solid #dc2626; }
.history-item-netral { border-left:5px solid #9ca3af; }
.history-badge-hadir { background:#dcfce7; color:#15803d; font-size:0.75rem; font-weight:600; padding:2px 10px; border-radius:999px; }
.history-badge-tidak { background:#fee2e2; color:#dc2626; font-size:0.75rem; font-weight:600; padding:2px 10px; border-radius:999px; }
.history-badge-netral { background:#f3f4f6; color:#6b7280; font-size:0.75rem; font-weight:600; padding:2px 10px; border-radius:999px; }
.stTabs [data-baseweb="tab-list"] { gap:6px; background:#f0fdf4; border-radius:12px; padding:5px; }
.stTabs [data-baseweb="tab"] { border-radius:8px; padding:8px 16px; font-weight:500; font-size:0.88rem; color:#166534; }
.stTabs [aria-selected="true"] { background-color:#14532d !important; color:white !important; }
.puskesmas-footer { background:rgba(20,83,45,0.06); padding:2.5rem 2rem 1.5rem; border-radius:20px 20px 0 0; margin-top:2rem; }
.footer-grid { display:grid; grid-template-columns:repeat(auto-fit,minmax(200px,1fr)); gap:2rem; margin-bottom:1.75rem; }
.footer-section h4 { font-size:0.95rem; font-weight:700; color:#14532d; border-bottom:1px solid rgba(20,83,45,0.1); padding-bottom:0.4rem; margin-bottom:0.75rem; }
.footer-section p,.footer-section a { font-size:0.85rem; color:#374151; line-height:1.9; text-decoration:none; }
.footer-section a:hover { color:#14532d; }
.footer-section i { width:18px; margin-right:6px; color:#166534; }
.footer-divider { border-top:1px solid rgba(0,0,0,0.05); margin:1rem 0; }
.footer-bottom { text-align:center; font-size:0.8rem; color:#6b7280; }
@media (max-width:768px) { .main-header h1 { font-size:1.6rem; } .footer-grid { grid-template-columns:1fr; } }
>>>>>>> f585c790da4d1a3b11f6ab84c1283f44dd5e6219
</style>
""", unsafe_allow_html=True)

<<<<<<< HEAD
API_BASE = "http://localhost:8000/api"

DAFTAR_NAMA = [
    "Isep Deni Herdian, S.Kep.,MMRS",
    "Isep Suhendar,SKM",
    "Bdn. Yeni Yulyani Setianingsih, S.ST",
    "Bdn. Nina Ainun, S.Tr.Keb",
    "Rita Sahara, S.Tr.Keb",
    "Dewi Sri Mulyani, Am.Keb",
    "Pipit Puspitasari, Am.Keb",
    "Mira Jatnikawati, Am.Keb",
    "Reni Mustikasari, Am.Keb",
    "Alitsa Nuur Fithri, S.ST",
    "Yesi Apriyani, Am.Keb",
    "Asri Awulan, S.Tr.Keb",
    "Pia Nur Podiana, A.Md.Keb",
    "Intang Sri Purnama, AM.Keb",
    "Ucu Lestari, AM.Keb",
    "Annisa Nafaulloh,S.Tr.Keb.,Bdn",
    "Mutia Wulansari.,S.Kep.,Ners",
    "Ujang Effendi, S.Kep.,Ners",
    "Liska Permatasari, S.Kep.,Ners",
    "Dede Khaerul Kamal Muchtar, AMK",
    "Iman Nurul Haq, A.Md.Kep",
    "Wida Idul Adha, S.Kep.,Ners",
    "Oriany Kemala Dewi, Amd.Kep",
    "Haeriah, A.Md.Kep",
    "Dede Aan Septiantini, A.Md.Kep",
    "dr.Ferry Nalapraya",
    "dr.Muhammad Azhary Romdhon",
    "dr.Iwan Setiawan",
    "dr. Siti Hana Fukui",
    "dr. Volti Diana Suryawadi",
    "drg.Rifan Hanggoro.M.M.R.S",
    "Endah Setiawati,S.Tr.Kes",
    "Khilman Husna Pratama, S.Farm.,Apt",
    "Vita Tyana Virista, A.Md.AK",
    "Gina Giovany, A.Md.AK",
    "Eko Wahyu Saputro, S.K.M",
    "Nurul Hasanah, A.Md.KL",
    "Nova Silpiany Perdany, A.Md.Farm",
    "Ameilia Putri Isyari, S.Gz",
    "Annisa Fauziah, A.Md.Gz",
    "Rudi Sutikno, SKM",
    "Yogi Aris Diyanto, S.E",
    "Rangga Ismardana Gasbela,S.T",
    "Winda Siti Sarah, AMd.RMIK",
    "Pupung Juliana",
    "Salsa Sabila",
    "Andina Dea Priatna, SKM",
    "Iip Supyan"
]

# ==================== DATA KARYAWAN BERDASARKAN ROLE ====================
def get_karyawan_by_role(role_keywords):
    """Filter karyawan berdasarkan role dari DAFTAR_NAMA"""
    role_map = {
        'dokter': ['dr.', 'drg.'],
        'perawat_ners': ['Ners', 'S.Kep', 'Amd.Kep', 'A.Md.Kep'],
        'bidan': ['Bdn.', 'S.Tr.Keb', 'Am.Keb', 'A.Md.Keb'],
        'promkes': ['Promosi', 'SKM'],
        'sanitarian': ['Sanitarian', 'S.K.M', 'A.Md.KL'],
        'gizi': ['S.Gz', 'A.Md.Gz'],
        'apoteker': ['Apt', 'S.Farm'],
        'lab': ['A.Md.AK'],
        'gigi': ['drg.', 'S.Tr.Kes'],
        'administrasi': ['S.E', 'S.T', 'S.Kep', 'S.ST', 'SKM', 'AMd.RMIK']
    }
    keywords = role_map.get(role_keywords, [])
    hasil = []
    for nama in DAFTAR_NAMA:
        for kw in keywords:
            if kw.lower() in nama.lower():
                hasil.append(nama)
                break
    return list(set(hasil))

# Karyawan tetap
PENDAFTARAN_TETAP = ["Winda Siti Sarah, AMd.RMIK", "Pupung Juliana", "Salsa Sabila"]
BP_GIGI_TETAP = ["drg.Rifan Hanggoro.M.M.R.S", "Endah Setiawati,S.Tr.Kes"]
APOTEK_TETAP = ["Khilman Husna Pratama, S.Farm.,Apt", "Nova Silpiany Perdany, A.Md.Farm"]
LAB_TETAP = ["Vita Tyana Virista, A.Md.AK", "Gina Giovany, A.Md.AK"]
PUSTU_CIANGIR = "Haeriah, A.Md.Kep"
PUSTU_SUMELAP = "Ujang Effendi, S.Kep.,Ners"
ADMINISTRASI_TETAP = ["Rangga Ismardana Gasbela,S.T", "Yogi Aris Diyanto, S.E"]
ADMINISTRASI_EXTRA = ["Liska Permatasari, S.Kep.,Ners", "Alitsa Nuur Fithri, S.ST", "Andina Dea Priatna, SKM"]

# ---------- Session State ----------
=======
>>>>>>> f585c790da4d1a3b11f6ab84c1283f44dd5e6219
for key, default in [
    ('logged_in', False), ('token', None), ('username', ''),
    ('page', 'user'), ('edit_data', None), ('edit_id', None),
    ('last_csv_url', ''), ('notif', None), ('pending_delete_ids', []),
    ('user_history_status', {}), ('show_csv_input', False),
    ('confirm_hapus_massal', False),
]:
    if key not in st.session_state:
        st.session_state[key] = default

API_BASE = "https://web-production-dc35a.up.railway.app/api"

DAFTAR_NAMA = [
    "Isep Deni Herdian, S.Kep.,MMRS","Isep Suhendar,SKM",
    "Bdn. Yeni Yulyani Setianingsih, S.ST","Bdn. Nina Ainun, S.Tr.Keb",
    "Rita Sahara, S.Tr.Keb","Dewi Sri Mulyani, Am.Keb",
    "Pipit Puspitasari, Am.Keb","Mira Jatnikawati, Am.Keb",
    "Reni Mustikasari, Am.Keb","Alitsa Nuur Fithri, S.ST",
    "Yesi Apriyani, Am.Keb","Asri Awulan, S.Tr.Keb",
    "Pia Nur Podiana, A.Md.Keb","Intang Sri Purnama, AM.Keb",
    "Ucu Lestari, AM.Keb","Annisa Nafaulloh,S.Tr.Keb.,Bdn",
    "Mutia Wulansari.,S.Kep.,Ners","Ujang Effendi, S.Kep.,Ners",
    "Liska Permatasari, S.Kep.,Ners","Dede Khaerul Kamal Muchtar, AMK",
    "Iman Nurul Haq, A.Md.Kep","Wida Idul Adha, S.Kep.,Ners",
    "Oriany Kemala Dewi, Amd.Kep","Haeriah, A.Md.Kep",
    "Dede Aan Septiantini, A.Md.Kep","dr.Ferry Nalapraya",
    "dr.Muhammad Azhary Romdhon","dr.Iwan Setiawan",
    "dr. Siti Hana Fukui","dr. Volti Diana Suryawadi",
    "drg.Rifan Hanggoro.M.M.R.S","Endah Setiawati,S.Tr.Kes",
    "Khilman Husna Pratama, S.Farm.,Apt","Vita Tyana Virista, A.Md.AK",
    "Gina Giovany, A.Md.AK","Eko Wahyu Saputro, S.K.M",
    "Nurul Hasanah, A.Md.KL","Nova Silpiany Perdany, A.Md.Farm",
    "Ameilia Putri Isyari, S.Gz","Annisa Fauziah, A.Md.Gz",
    "Rudi Sutikno, SKM","Yogi Aris Diyanto, S.E",
    "Rangga Ismardana Gasbela,S.T","Winda Siti Sarah, AMd.RMIK",
    "Pupung Juliana","Salsa Sabila","Andina Dea Priatna, SKM","Iip Supyan"
]

PENDAFTARAN_TETAP  = ["Winda Siti Sarah, AMd.RMIK","Pupung Juliana","Salsa Sabila"]
BP_GIGI_TETAP      = ["drg.Rifan Hanggoro.M.M.R.S","Endah Setiawati,S.Tr.Kes"]
APOTEK_TETAP       = ["Khilman Husna Pratama, S.Farm.,Apt","Nova Silpiany Perdany, A.Md.Farm"]
LAB_TETAP          = ["Vita Tyana Virista, A.Md.AK","Gina Giovany, A.Md.AK"]
PUSTU_CIANGIR      = "Haeriah, A.Md.Kep"
PUSTU_SUMELAP      = "Ujang Effendi, S.Kep.,Ners"
ADMINISTRASI_TETAP = ["Rangga Ismardana Gasbela,S.T","Yogi Aris Diyanto, S.E"]
ADMINISTRASI_EXTRA = ["Liska Permatasari, S.Kep.,Ners","Alitsa Nuur Fithri, S.ST","Andina Dea Priatna, SKM"]
NAMA_DIECUALIKAN   = ["Isep Deni Herdian, S.Kep.,MMRS","Isep Suhendar,SKM"]

ROLE_MAP = {
    'dokter':['dr.','drg.'],
    'perawat_ners':['Ners','S.Kep','Amd.Kep','A.Md.Kep'],
    'bidan':['Bdn.','S.Tr.Keb','Am.Keb','A.Md.Keb'],
    'promkes':['Promosi','SKM'],
    'sanitarian':['Sanitarian','S.K.M','A.Md.KL'],
    'gizi':['S.Gz','A.Md.Gz'],
    'apoteker':['Apt','S.Farm'],
    'lab':['A.Md.AK'],
    'gigi':['drg.','S.Tr.Kes'],
    'administrasi':['S.E','S.T','S.Kep','S.ST','SKM','AMd.RMIK'],
}

def get_karyawan_by_role(role_key, exclude=None):
    if exclude is None: exclude = []
    keywords = ROLE_MAP.get(role_key, [])
    hasil = []
    for nama in DAFTAR_NAMA:
        if nama in exclude: continue
        if any(kw.lower() in nama.lower() for kw in keywords):
            hasil.append(nama)
    return list(set(hasil))

def filter_nama(nama_list):
    return [n for n in nama_list if n not in NAMA_DIECUALIKAN]

def auth_headers():
    return {"Authorization": f"Token {st.session_state.token}"}

def api_get(path, params=None, auth=False):
    h = auth_headers() if auth else {}
    return requests.get(f"{API_BASE}/{path}", params=params, headers=h, timeout=10)

def api_post(path, data, auth=False):
    h = {"Content-Type": "application/json"}
    if auth: h.update(auth_headers())
    return requests.post(f"{API_BASE}/{path}", json=data, headers=h, timeout=10)

def api_put(path, data):
    return requests.put(f"{API_BASE}/{path}", json=data, headers=auth_headers(), timeout=10)

def api_delete(path):
    return requests.delete(f"{API_BASE}/{path}", headers=auth_headers(), timeout=10)

def parse_penyerta(teks):
<<<<<<< HEAD
    if not teks:
        return []
=======
    if not teks: return []
>>>>>>> f585c790da4d1a3b11f6ab84c1283f44dd5e6219
    if ';' in teks:
        return [p.strip() for p in teks.split(';') if p.strip()]
    parts = re.split(r', (?=[A-Z])', teks)
    return [p.strip() for p in parts if p.strip()]

<<<<<<< HEAD
# ==================== FUNGSI RANDOMIZE JADWAL ====================
def random_pick_from_list(lst, count=1, exclude=None):
    if exclude is None:
        exclude = []
    available = [x for x in lst if x not in exclude]
    if len(available) < count:
        return available
    return random.sample(available, count)

def generate_jadwal_bulanan(month, year):
    
    """Generate jadwal untuk satu bulan penuh (Senin-Sabtu)"""
    random.seed(f"{year}-{month}")
    
    cal = calendar.monthcalendar(year, month)
    work_days = []
    for week in cal:
        for day_idx, day in enumerate(week):
            if day != 0 and day_idx < 6:
                work_days.append(day)
    
    day_names = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu"]
    
    semua_dokter = get_karyawan_by_role('dokter')
    semua_perawat_ners = get_karyawan_by_role('perawat_ners')
    semua_bidan = get_karyawan_by_role('bidan')
    semua_promkes = get_karyawan_by_role('promkes')
    semua_sanitarian = get_karyawan_by_role('sanitarian')
    semua_gizi = get_karyawan_by_role('gizi')
    
    pool_ilp_prolanis = list(set(semua_perawat_ners + semua_bidan + semua_promkes + semua_sanitarian + semua_gizi))
    
    jadwal_generated = []
    
    for tanggal in work_days:
        tgl_str = f"{year}-{month:02d}-{tanggal:02d}"
        tgl_obj = datetime(year, month, tanggal)
        hari_ke = tgl_obj.weekday()
        nama_hari = day_names[hari_ke]
        
        used_today = set()
        jadwal_hari = []
        
        # 1. PENDAFTARAN
        jadwal_hari.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': 'PENDAFTARAN', 'penyerta': '; '.join(PENDAFTARAN_TETAP)})
        used_today.update(PENDAFTARAN_TETAP)
        
        # 2. SKRINING ILP 1 & 2
        for i in range(1, 3):
            available = [p for p in pool_ilp_prolanis if p not in used_today]
            if available:
                petugas = random.choice(available)
                used_today.add(petugas)
                jadwal_hari.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': f'SKRINING ILP {i}', 'penyerta': petugas})
        
        # 3. POLI PROLANIS
        available_prolanis = [p for p in pool_ilp_prolanis if p not in used_today]
        if len(available_prolanis) >= 3:
            petugas_prolanis = random.sample(available_prolanis, 3)
            used_today.update(petugas_prolanis)
            jadwal_hari.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': 'POLI PROLANIS', 'penyerta': '; '.join(petugas_prolanis)})
        
        # 4. KLASTER DEWASA-LANSIA 1 & 2
        for i in range(1, 3):
            available_dokter = [d for d in semua_dokter if d not in used_today]
            if available_dokter:
                petugas = random.choice(available_dokter)
                used_today.add(petugas)
            else:
                available_ners = [n for n in semua_perawat_ners if n not in used_today]
                petugas = random.choice(available_ners) if available_ners else "Tidak tersedia"
                if petugas != "Tidak tersedia":
                    used_today.add(petugas)
            jadwal_hari.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': f'KLASTER DEWASA-LANSIA {i}', 'penyerta': petugas})
        
        # 5. KLASTER IBU KIA & USG
        available_bidan = [b for b in semua_bidan if b not in used_today]
        available_dokter = [d for d in semua_dokter if d not in used_today]
        petugas_bidan = random.sample(available_bidan, 2) if len(available_bidan) >= 2 else available_bidan
        used_today.update(petugas_bidan)
        petugas_dokter = random.choice(available_dokter) if available_dokter else "Tidak tersedia"
        if petugas_dokter != "Tidak tersedia":
            used_today.add(petugas_dokter)
        jadwal_hari.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': 'KLASTER IBU KIA & USG', 'penyerta': f"{'; '.join(petugas_bidan)}; {petugas_dokter}"})
        
        # 6. KLASTER ANAK
        available_bidan = [b for b in semua_bidan if b not in used_today]
        available_dokter = [d for d in semua_dokter if d not in used_today]
        petugas_bidan = random.sample(available_bidan, 2) if len(available_bidan) >= 2 else available_bidan
        used_today.update(petugas_bidan)
        petugas_dokter = random.choice(available_dokter) if available_dokter else "Tidak tersedia"
        if petugas_dokter != "Tidak tersedia":
            used_today.add(petugas_dokter)
        jadwal_hari.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': 'KLASTER ANAK', 'penyerta': f"{'; '.join(petugas_bidan)}; {petugas_dokter}"})
        
        # 7. R. IMUNISASI (hanya Kamis)
        if nama_hari == "Kamis":
            available_bidan = [b for b in semua_bidan if b not in used_today]
            petugas_imunisasi = random.sample(available_bidan, 2) if len(available_bidan) >= 2 else available_bidan
            used_today.update(petugas_imunisasi)
            jadwal_hari.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': 'R. IMUNISASI', 'penyerta': '; '.join(petugas_imunisasi)})
        
        # 8. R. TINDAKAN
        available_perawat = [p for p in semua_perawat_ners if p not in used_today]
        petugas_tindakan = random.choice(available_perawat) if available_perawat else "Tidak tersedia"
        if petugas_tindakan != "Tidak tersedia":
            used_today.add(petugas_tindakan)
        jadwal_hari.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': 'R. TINDAKAN', 'penyerta': petugas_tindakan})
        
        # 9. BP GIGI
        jadwal_hari.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': 'BP GIGI', 'penyerta': '; '.join(BP_GIGI_TETAP)})
        used_today.update(BP_GIGI_TETAP)
        
        # 10. APOTEK
        jadwal_hari.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': 'APOTEK', 'penyerta': '; '.join(APOTEK_TETAP)})
        used_today.update(APOTEK_TETAP)
        
        # 11. LAB
        jadwal_hari.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': 'LAB', 'penyerta': '; '.join(LAB_TETAP)})
        used_today.update(LAB_TETAP)
        
        # 12. R. TB (hanya Selasa)
        if nama_hari == "Selasa":
            available_perawat = [p for p in semua_perawat_ners if p not in used_today]
            petugas_tb = random.choice(available_perawat) if available_perawat else "Tidak tersedia"
            if petugas_tb != "Tidak tersedia":
                used_today.add(petugas_tb)
            jadwal_hari.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': 'R. TB', 'penyerta': petugas_tb})
        
        # 13. ADMINISTRASI
        admin_extra = random_pick_from_list(ADMINISTRASI_EXTRA, count=2)
        semua_admin = ADMINISTRASI_TETAP + admin_extra
        jadwal_hari.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': 'ADMINISTRASI', 'penyerta': '; '.join(semua_admin)})
        used_today.update(semua_admin)
        
        # 14. PUSTU CIANGIR
        jadwal_hari.append({'tanggal': tgl_str, 'lokasi': 'Pustu Ciangir', 'kegiatan': 'PELAYANAN PUSTU', 'penyerta': PUSTU_CIANGIR})
        used_today.add(PUSTU_CIANGIR)
        
        # 15. PUSTU SUMELAP
        jadwal_hari.append({'tanggal': tgl_str, 'lokasi': 'Pustu Sumelap', 'kegiatan': 'PELAYANAN PUSTU', 'penyerta': PUSTU_SUMELAP})
        used_today.add(PUSTU_SUMELAP)
        
        # 16. PIKET PERSALINAN (bidan yang belum dipakai)
        all_bidan = set(semua_bidan)
        used_bidan = used_today.intersection(all_bidan)
        available_bidan_piket = list(all_bidan - used_bidan)
        if available_bidan_piket:
            petugas_piket = random.choice(available_bidan_piket)
            jadwal_hari.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': 'PIKET PERSALINAN', 'penyerta': petugas_piket})
        
        jadwal_generated.extend(jadwal_hari)
    
    return jadwal_generated
# ==================== FUNGSI PDF (VERSI DIPERBAIKI) ====================
def generate_jadwal_pdf(data):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                            leftMargin=0.3*inch, rightMargin=0.3*inch,
                            topMargin=0.4*inch, bottomMargin=0.4*inch)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=16, alignment=1, spaceAfter=20)
    
    story = []
    story.append(Paragraph("<b>JADWAL PELAYANAN PUSKESMAS SANGKALI</b>", title_style))
    story.append(Paragraph(f"SEMUA JADWAL - {datetime.now().strftime('%d %B %Y')}", styles['Heading2']))
    story.append(Spacer(1, 15))
    
    df = pd.DataFrame(data)
    
    # PERBAIKAN: Pastikan kolom tanggal dalam format datetime
    if 'tanggal' in df.columns:
        df['tanggal'] = pd.to_datetime(df['tanggal'], errors='coerce').dt.date
    
    days = sorted(df['tanggal'].dropna().unique())
    if len(days) == 0:
        story.append(Paragraph("Tidak ada data jadwal yang ditemukan.", styles['Normal']))
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    day_headers = [d.strftime('%A, %d\n%b') for d in days]
    
    ruang_kegiatan = [
        'PENDAFTARAN', 'SKRINING ILP 1', 'SKRINING ILP 2', 'POLI PROLANIS',
        'KLASTER DEWASA-LANSIA 1', 'KLASTER DEWASA-LANSIA 2', 'KLASTER IBU KIA & USG',
        'KLASTER ANAK', 'R. IMUNISASI', 'R. TINDAKAN', 'BP GIGI', 'APOTEK', 'LAB',
        'R. TB', 'ADMINISTRASI', 'PUSTU CIANGIR', 'PUSTU SUMELAP'
    ]
    
    table_data = [['RUANG PELAYANAN'] + day_headers]
    
    for keg in ruang_kegiatan:
        row = [keg]
        for d in days:
            # Perbaikan matching tanggal
            items = df[df['tanggal'] == d]
            items = items[items['kegiatan'].str.contains(keg, na=False, case=False)]
            
            if not items.empty:
                peny = items.iloc[0]['penyerta']
                row.append(peny[:50] + '...' if len(peny) > 50 else peny)
            else:
                row.append('')
        table_data.append(row)
    
    t = Table(table_data, colWidths=[2.2*inch] + [1.35*inch] * len(days))
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.darkgreen),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('BACKGROUND', (0,1), (-1,-1), colors.white),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(t)
    story.append(Spacer(1, 30))
    story.append(Paragraph("Dicetak dari Sistem Puskesmas Sangkali © 2026", styles['Normal']))
    
=======
def show_notif(message, notif_type="success"):
    icons = {"success":"✅","error":"❌","info":"ℹ️"}
    st.toast(message, icon=icons.get(notif_type,"ℹ️"))

# ── PARSE CSV GOOGLE SHEET ──
def parse_google_sheet_csv(raw_text):
    import csv as _csv
    lines = raw_text.strip().splitlines()
    if not lines:
        return []

    header_parts = [h.strip().lower() for h in lines[0].split(',')]
    col_map = {}
    for i, h in enumerate(header_parts):
        if 'tanggal' in h or h == 'date':
            col_map['tanggal'] = i
        elif 'lokasi' in h or h in ('location', 'tempat'):
            col_map['lokasi'] = i
        elif 'kegiatan' in h or 'activity' in h or 'nama kegiatan' in h:
            col_map['kegiatan'] = i
        elif 'penyerta' in h or 'pelaksana' in h or 'peserta' in h or 'petugas' in h:
            col_map['penyerta'] = i

    required = ['tanggal', 'lokasi', 'kegiatan', 'penyerta']
    missing = [k for k in required if k not in col_map]
    if missing:
        raise ValueError(f"Kolom tidak ditemukan: {missing}. Header terdeteksi: {header_parts}")

    peny_idx = col_map['penyerta']
    results = []

    for line in lines[1:]:
        line = line.strip()
        if not line:
            continue

        try:
            row = list(_csv.reader([line]))[0]
        except Exception:
            row = []

        if len(row) >= 4:
            if len(row) > peny_idx + 1:
                penyerta_val = ', '.join(r.strip() for r in row[peny_idx:])
            else:
                penyerta_val = row[peny_idx].strip()
            record = {
                'tanggal':  row[col_map['tanggal']].strip(),
                'lokasi':   row[col_map['lokasi']].strip(),
                'kegiatan': row[col_map['kegiatan']].strip(),
                'penyerta': penyerta_val,
            }
        else:
            parts = line.split(',', peny_idx)
            if len(parts) <= peny_idx:
                continue
            record = {
                'tanggal':  parts[col_map.get('tanggal',  0)].strip(),
                'lokasi':   parts[col_map.get('lokasi',   1)].strip(),
                'kegiatan': parts[col_map.get('kegiatan', 2)].strip(),
                'penyerta': parts[peny_idx].strip().strip('"'),
            }

        if record['tanggal'] and record['kegiatan']:
            results.append(record)

    return results


def sync_from_url(csv_url, mode, auth=True):
    try:
        resp = requests.get(csv_url, timeout=15)
        resp.raise_for_status()
        raw = resp.text
    except Exception as e:
        raise ValueError(f"Gagal mengunduh CSV: {e}")

    records = parse_google_sheet_csv(raw)
    if not records:
        raise ValueError("Tidak ada data valid di spreadsheet.")

    if mode == 'replace':
        try:
            r_all = api_get("kegiatan/", auth=True)
            if r_all.status_code == 200:
                ids_all = [item['id'] for item in r_all.json()]
                if ids_all:
                    api_post("kegiatan/bulk-delete/", {"ids": ids_all}, auth=True)
        except Exception:
            pass

    existing_set = set()
    if mode == 'append':
        try:
            r_ex = api_get("kegiatan/", auth=True)
            if r_ex.status_code == 200:
                for item in r_ex.json():
                    existing_set.add((item['tanggal'], item['lokasi'], item['kegiatan']))
        except Exception:
            pass

    saved = skipped = errors = 0
    for rec in records:
        key = (rec['tanggal'], rec['lokasi'], rec['kegiatan'])
        if mode == 'append' and key in existing_set:
            skipped += 1
            continue
        try:
            r = api_post("kegiatan/", rec, auth=True)
            if r.status_code == 201:
                saved += 1
                existing_set.add(key)
            else:
                errors += 1
        except Exception:
            errors += 1

    return saved, skipped, errors

# ── PDF ──
def generate_jadwal_pdf(data):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                            leftMargin=0.4*inch, rightMargin=0.4*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CT', parent=styles['Title'], fontSize=14, alignment=1,
                                 spaceAfter=6, textColor=colors.HexColor('#14532d'))
    sub_style   = ParagraphStyle('CS', parent=styles['Normal'], fontSize=10, alignment=1,
                                 spaceAfter=14, textColor=colors.HexColor('#166534'))
    cell_style  = ParagraphStyle('CE', parent=styles['Normal'], fontSize=7, leading=9, wordWrap='LTR')
    foot_style  = ParagraphStyle('CF', parent=styles['Normal'], fontSize=8,
                                 textColor=colors.grey, alignment=1)

    story = [
        Paragraph("JADWAL PELAYANAN PUSKESMAS SANGKALI", title_style),
        Paragraph(f"Dicetak: {datetime.now().strftime('%d %B %Y, %H:%M')}", sub_style),
        Spacer(1, 6)
    ]

    if not data:
        story.append(Paragraph("Tidak ada data jadwal.", styles['Normal']))
        doc.build(story); buffer.seek(0); return buffer

    df = pd.DataFrame(data)
    for col in ['tanggal','lokasi','kegiatan','penyerta']:
        if col not in df.columns: df[col] = ''
    df['tanggal'] = pd.to_datetime(df['tanggal'], errors='coerce')
    df = df.sort_values('tanggal')
    df['tanggal_str'] = df['tanggal'].dt.strftime('%A,\n%d %b %Y')

    col_widths = [1.5*inch, 1.4*inch, 2.0*inch, 5.6*inch]
    headers    = ['TANGGAL','LOKASI','KEGIATAN','PELAKSANA / PENYERTA']
    table_data = [[Paragraph(f'<b>{h}</b>', cell_style) for h in headers]]
    for _, row in df.iterrows():
        tgl_disp = row['tanggal_str'] if pd.notna(row['tanggal']) else str(row.get('tanggal',''))
        peny_raw = str(row.get('penyerta','')).replace(';','\n').strip()
        table_data.append([
            Paragraph(tgl_disp, cell_style),
            Paragraph(str(row.get('lokasi','')), cell_style),
            Paragraph(str(row.get('kegiatan','')), cell_style),
            Paragraph(peny_raw, cell_style),
        ])

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#14532d')),
        ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
        ('ALIGN',(0,0),(-1,-1),'LEFT'),
        ('VALIGN',(0,0),(-1,-1),'TOP'),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,0),8),
        ('BOTTOMPADDING',(0,0),(-1,0),8),
        ('TOPPADDING',(0,0),(-1,0),8),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#f0fdf4')]),
        ('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#d1d5db')),
        ('TOPPADDING',(0,1),(-1,-1),5),
        ('BOTTOMPADDING',(0,1),(-1,-1),5),
        ('LEFTPADDING',(0,0),(-1,-1),6),
        ('RIGHTPADDING',(0,0),(-1,-1),6),
    ]))
    story.append(t)
    story.append(Spacer(1,20))
    story.append(Paragraph(
        f"Dicetak dari Sistem Puskesmas Sangkali © {datetime.now().year} | Total: {len(df)} kegiatan",
        foot_style
    ))
>>>>>>> f585c790da4d1a3b11f6ab84c1283f44dd5e6219
    doc.build(story)
    buffer.seek(0)
    return buffer

<<<<<<< HEAD
=======
# ── CSV ──
def generate_csv(data, filename="jadwal.csv"):
    if not data:
        st.download_button("Download CSV", data=b"TANGGAL,LOKASI,KEGIATAN,PELAKSANA\n",
                           file_name=filename, mime='text/csv', use_container_width=True)
        return
    df = pd.DataFrame(data)
    rename_map = {'tanggal':'TANGGAL','lokasi':'LOKASI','kegiatan':'KEGIATAN','penyerta':'PELAKSANA'}
    df = df.rename(columns={k:v for k,v in rename_map.items() if k in df.columns})
    kolom = [c for c in ['TANGGAL','LOKASI','KEGIATAN','PELAKSANA'] if c in df.columns]
    df = df[kolom]
    if 'TANGGAL' in df.columns:
        df['TANGGAL'] = pd.to_datetime(df['TANGGAL'], errors='coerce').dt.strftime('%Y-%m-%d')
    buf = StringIO()
    df.to_csv(buf, index=False, sep=',', quoting=1)
    csv_bytes = buf.getvalue().encode('utf-8-sig')
    st.download_button("⬇ Download CSV", data=csv_bytes, file_name=filename,
                       mime='text/csv', use_container_width=True)

# ── TOKEN URL ──
params = st.query_params
if not st.session_state.logged_in and 'token' in params:
    token = params['token']
    try:
        resp = requests.get(f"{API_BASE}/verify-token/", headers={"Authorization":f"Token {token}"}, timeout=5)
        if resp.status_code == 200:
            d = resp.json()
            st.session_state.logged_in = True
            st.session_state.token = token
            st.session_state.username = d.get('username','Admin')
            st.query_params.clear(); st.rerun()
        else: st.query_params.clear()
    except: pass

>>>>>>> f585c790da4d1a3b11f6ab84c1283f44dd5e6219
if st.session_state.notif:
    show_notif(st.session_state.notif)
    st.session_state.notif = None

# ── SIDEBAR ──
with st.sidebar:
    cl, cc, cr = st.columns([0.5,3,0.5])
    with cc:
        try: st.image("assets/logo.png", width=90)
        except:
            st.markdown("""<div style="display:flex;justify-content:center;">
            <svg width="80" height="80" viewBox="0 0 80 80">
            <circle cx="40" cy="40" r="38" fill="rgba(255,255,255,0.15)" stroke="rgba(255,255,255,0.3)" stroke-width="2"/>
            <rect x="35" y="20" width="10" height="40" rx="3" fill="white"/>
            <rect x="20" y="35" width="40" height="10" rx="3" fill="white"/>
            </svg></div>""", unsafe_allow_html=True)
    st.markdown('<div class="sidebar-puskesmas-name">UPTD Puskesmas<br>Sangkali</div>', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-puskesmas-sub">Tasikmalaya, Jawa Barat</div>', unsafe_allow_html=True)
    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown('<div class="sidebar-menu-label">Navigasi</div>', unsafe_allow_html=True)
    if st.button("Jadwal Kegiatan", use_container_width=True):
        st.session_state.page="user"; st.rerun()
    if not st.session_state.logged_in:
        if st.button("Login Admin", use_container_width=True):
            st.session_state.page="admin_login"; st.rerun()
    else:
        st.markdown(f'<div class="sidebar-user-badge"><i class="fa-solid fa-circle-user"></i> {st.session_state.username}</div>', unsafe_allow_html=True)
        if st.button("Dashboard Admin", use_container_width=True):
            st.session_state.page="admin_dashboard"; st.rerun()
        if st.button("Logout", use_container_width=True):
            try: api_post("logout/",{},auth=True)
            except: pass
            for k in ['logged_in','token','username']:
                st.session_state[k] = False if k=='logged_in' else None if k=='token' else ''
            st.session_state.page="user"; st.query_params.clear(); st.rerun()
    st.markdown("<hr>", unsafe_allow_html=True)
    try:
        r = requests.get(f"{API_BASE}/jadwal-terdekat/", timeout=3)
        if r.status_code==200: st.markdown('<span class="badge-online"><i class="fa-solid fa-circle-dot"></i> Online</span>', unsafe_allow_html=True)
        else: st.markdown('<span class="badge-offline"><i class="fa-solid fa-circle-exclamation"></i> Error</span>', unsafe_allow_html=True)
    except: st.markdown('<span class="badge-offline"><i class="fa-solid fa-circle-exclamation"></i> Offline</span>', unsafe_allow_html=True)
    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown('<div style="font-size:0.7rem;opacity:0.6;text-align:center;"><i class="fa-solid fa-clock"></i> Senin–Jumat 07:30–14:00</div>', unsafe_allow_html=True)

<<<<<<< HEAD
# ========================== HALAMAN USER ==========================
if st.session_state.page == "user":
=======
# ── HEADER ──
if st.session_state.page=="user":
    st.markdown('<div class="main-header"><h1>Puskesmas Sangkali</h1><p>Sistem Informasi Manajemen Kegiatan</p></div>', unsafe_allow_html=True)
else:
    st.markdown('<div class="main-header"><h1>Puskesmas Sangkali</h1><p>Dashboard Manajemen</p></div>', unsafe_allow_html=True)

# ════════════════════════════════════════
# USER PAGE
# ════════════════════════════════════════
if st.session_state.page=="user":
>>>>>>> f585c790da4d1a3b11f6ab84c1283f44dd5e6219
    try:
        today=date.today(); tomorrow=today+timedelta(days=1)
        resp_notif=requests.get(f"{API_BASE}/jadwal-terdekat/", timeout=3)
        if resp_notif.status_code==200:
            jadwal=resp_notif.json()
            notif_hari_ini=[j for j in jadwal if j['tanggal']==str(today)]
            notif_besok=[j for j in jadwal if j['tanggal']==str(tomorrow)]
            if notif_hari_ini:
                items=''.join(f"<li><b>{j['kegiatan']}</b> — {j['lokasi']}</li>" for j in notif_hari_ini[:6])
                st.markdown(f"<div style='background:#dcfce7;border:1px solid #bbf7d0;padding:1rem;border-radius:12px;margin-bottom:0.75rem;'><i class='fa-solid fa-circle-exclamation'></i> <b>Hari ini:</b><ul>{items}</ul></div>", unsafe_allow_html=True)
            if notif_besok:
                items=''.join(f"<li><b>{j['kegiatan']}</b> — {j['lokasi']}</li>" for j in notif_besok[:6])
                st.markdown(f"<div style='background:#eff6ff;border:1px solid #bfdbfe;padding:1rem;border-radius:12px;margin-bottom:0.75rem;'><i class='fa-solid fa-calendar-check'></i> <b>Besok:</b><ul>{items}</ul></div>", unsafe_allow_html=True)
    except: pass

    col1,col2=st.columns([2,1])
    with col1:
        st.markdown('<div style="display:flex;align-items:center;gap:8px;margin-bottom:0.5rem"><i class="fa-solid fa-magnifying-glass" style="color:#14532d;"></i><span style="font-size:1.3rem;font-weight:700;color:#14532d;">Cari Jadwal Kegiatan</span></div>', unsafe_allow_html=True)
        with st.form("user_form"):
<<<<<<< HEAD
            nama = st.selectbox("Nama Lengkap", DAFTAR_NAMA)
            tanggal = st.date_input("Tanggal Kegiatan", datetime.now().date())
            if st.form_submit_button("Cari Jadwal"):
                try:
                    resp = api_get("search-user/", params={"nama": nama, "tanggal": str(tanggal)})
                    if resp.status_code == 200:
                        hasil = resp.json()
                        if hasil:
                            st.success(f"Halo {nama}")
                            for h in hasil:
                                peny_list = parse_penyerta(h['penyerta'])
                                peny_items = ''.join(f"<li>{p}</li>" for p in peny_list)
=======
            nama=st.selectbox("Nama Lengkap", DAFTAR_NAMA)
            tanggal=st.date_input("Tanggal", datetime.now().date())
            if st.form_submit_button("Cari"):
                resp=api_get("search-user/", params={"nama":nama,"tanggal":str(tanggal)})
                if resp.status_code==200 and resp.json():
                    hasil=resp.json()
                    tgl_fmt=tanggal.strftime('%A, %d %B %Y')
                    items=""
                    for h in hasil:
                        peny_list = parse_penyerta(h['penyerta'])
                        peny_items = ''.join(f'<span>{p}</span>' for p in peny_list) if peny_list else f'<span>{h["penyerta"]}</span>'
                        items+=f"""<div class="receipt-item">
                            <div class="receipt-item-title">{h['kegiatan']}</div>
                            <div class="receipt-item-row"><span class="receipt-item-label">Lokasi</span><span class="receipt-item-value">{h['lokasi']}</span></div>
                            <div class="receipt-item-row"><span class="receipt-item-label">Penyerta</span><div class="receipt-item-value"><div class="receipt-penyerta-list">{peny_items}</div></div></div>
                        </div>"""
                    st.markdown(f"<div class='receipt-box'><div class='receipt-header'><h3>Jadwal Kegiatan</h3><p>{tgl_fmt}</p></div><div class='receipt-body'>{items}</div><div class='receipt-footer'><div class='receipt-barcode'>|||||||||||||||||||||||</div>Puskesmas Sangkali</div></div>", unsafe_allow_html=True)
                else: st.warning("Tidak ada kegiatan.")
    with col2:
        st.markdown('<div class="data-card"><div style="display:flex;align-items:center;gap:6px;margin-bottom:0.5rem"><i class="fa-solid fa-circle-info" style="color:#14532d;"></i><span style="font-weight:700;color:#14532d;">Informasi</span></div><ul style="font-size:0.88rem;line-height:1.8;"><li>Datang 15 menit lebih awal</li><li>Bawa KMS/BPJS</li><li>Gunakan masker</li></ul></div>', unsafe_allow_html=True)

        # ── Riwayat Kehadiran ──
    st.markdown("---")
    st.markdown('<div style="display:flex;align-items:center;gap:8px;margin-bottom:0.5rem"><i class="fa-solid fa-clock-rotate-left" style="color:#14532d;"></i><span style="font-size:1.2rem;font-weight:700;color:#14532d;">Riwayat Kehadiran Saya</span></div>', unsafe_allow_html=True)
    cf1, cf2, cf3 = st.columns([2, 2, 1])
    with cf1:
        nama_riwayat = st.selectbox("Nama", DAFTAR_NAMA, key="nama_riwayat")
    with cf2:
        bulan_riwayat = st.selectbox("Bulan", ["Semua"] + list(calendar.month_name)[1:], key="bulan_riwayat")
    with cf3:
        tahun_riwayat = st.number_input("Tahun", value=datetime.now().year, min_value=2020, max_value=2030)

    if st.button("Tampilkan Riwayat"):
        try:
            resp = api_get("search-user/", params={"nama": nama_riwayat})
            raw_data = None
            if resp.status_code == 200:
                raw_data = resp.json()
            else:
                for use_auth in ([True, False] if st.session_state.logged_in else [False]):
                    try:
                        r2 = api_get("kegiatan/", auth=use_auth)
                        if r2.status_code == 200:
                            all_d = r2.json()
                            kw = nama_riwayat.split(",")[0].strip()
                            raw_data = [d for d in all_d if kw.lower() in d.get("penyerta", "").lower()]
                            break
                    except:
                        continue

            if not raw_data:
                st.info("Belum ada data kegiatan di sistem.")
            else:
                df = pd.DataFrame(raw_data)
                df["tanggal_dt"] = pd.to_datetime(df["tanggal"], errors="coerce")
                df_filtered = df[df["tanggal_dt"].dt.year == tahun_riwayat]
                if bulan_riwayat != "Semua":
                    bulan_num = list(calendar.month_name).index(bulan_riwayat)
                    df_filtered = df_filtered[df_filtered["tanggal_dt"].dt.month == bulan_num]
                df_filtered = df_filtered.sort_values("tanggal_dt", ascending=False)

                if df_filtered.empty:
                    st.info(f"Tidak ada kegiatan ditemukan untuk {nama_riwayat}.")
                else:
                    # Header receipt
                    st.markdown(f"""
                    <div class="receipt-box" style="margin-bottom:0;border-radius:4px 4px 0 0;">
                        <div class="receipt-header">
                            <h3>Riwayat Kehadiran</h3>
                            <p>{nama_riwayat}</p>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

                    today_ = date.today()
                    rows_list = list(df_filtered.iterrows())

                    for idx, (_, row) in enumerate(rows_list):
                        unique_key = f"{row['tanggal']}|{row['kegiatan']}|{row['lokasi']}"
                        status = st.session_state.user_history_status.get(unique_key, "belum")
                        tgl_dt = row["tanggal_dt"]
                        tgl_fmt = tgl_dt.strftime("%A, %d %B %Y")
                        sudah_lewat = tgl_dt.date() <= today_
                        is_last = (idx == len(rows_list) - 1)

                        # Badge status kehadiran
                        if status == "hadir":
                            badge_html = '<span class="history-badge-hadir">Hadir</span>'
                        elif status == "tidak_hadir":
                            badge_html = '<span class="history-badge-tidak">Tidak Hadir</span>'
                        else:
                            badge_html = '<span class="history-badge-netral">Belum Ditandai</span>'

                        # Penyerta — satu baris per nama
                        peny_list = parse_penyerta(str(row["penyerta"]))
                        peny_html = "".join(
                            f'<span style="display:block;margin-bottom:0.1rem;">{p}</span>'
                            for p in peny_list
                        ) if peny_list else f'<span>{str(row["penyerta"])}</span>'

                        border_bottom = "border-bottom:1px dashed #ddd8cc;" if not is_last else ""

                        # Kolom: receipt info (kiri) | tombol (kanan, hanya jika sudah lewat)
                        if sudah_lewat:
                            col_info, col_btn = st.columns([4, 1])
                        else:
                            col_info = st.columns(1)[0]
                            col_btn = None

                        with col_info:
                            st.markdown(f"""
                            <div style="background:#fffef8;
                                        border-left:1px solid #e2e0d8;
                                        border-right:{'none' if sudah_lewat else '1px solid #e2e0d8'};
                                        padding:0.75rem 1.25rem;
                                        {border_bottom}">
                                <div class="receipt-item-title" style="margin-bottom:0.3rem;">{row['kegiatan']}</div>
                                <div class="receipt-item-row">
                                    <span class="receipt-item-label">Tanggal</span>
                                    <span class="receipt-item-value">{tgl_fmt}</span>
                                </div>
                                <div class="receipt-item-row">
                                    <span class="receipt-item-label">Lokasi</span>
                                    <span class="receipt-item-value">{row['lokasi']}</span>
                                </div>
                                <div class="receipt-item-row" style="align-items:flex-start;">
                                    <span class="receipt-item-label">Penyerta</span>
                                    <div class="receipt-item-value">{peny_html}</div>
                                </div>
                                <div style="margin-top:0.35rem;">{badge_html}</div>
                            </div>
                            """, unsafe_allow_html=True)

                        if col_btn is not None:
                            with col_btn:
>>>>>>> f585c790da4d1a3b11f6ab84c1283f44dd5e6219
                                st.markdown(f"""
                                <div style="background:#fffef8;
                                            border-right:1px solid #e2e0d8;
                                            padding:0.75rem 0.5rem;
                                            height:100%;
                                            display:flex;
                                            flex-direction:column;
                                            justify-content:center;
                                            gap:6px;
                                            {border_bottom}">
                                </div>
                                """, unsafe_allow_html=True)
                                if st.button(
                                    "✓ Hadir",
                                    key=f"hadir_{unique_key}_{idx}",
                                    use_container_width=True,
                                    type="primary" if status != "hadir" else "secondary",
                                ):
                                    st.session_state.user_history_status[unique_key] = "hadir"
                                    st.rerun()
                                if st.button(
                                    "✗ Tidak",
                                    key=f"tidak_{unique_key}_{idx}",
                                    use_container_width=True,
                                ):
                                    st.session_state.user_history_status[unique_key] = "tidak_hadir"
                                    st.rerun()

                    # Footer receipt
                    st.markdown(f"""
                    <div class="receipt-box" style="margin-top:0;border-radius:0 0 4px 4px;
                         border-top:none;box-shadow:none;">
                        <div class="receipt-footer">
                            <div class="receipt-barcode">|||||||||||||||||||||||</div>
                            {len(df_filtered)} kegiatan
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

<<<<<<< HEAD
    st.markdown("""
    <div class="icon-text" style="margin-top:1.5rem; margin-bottom:0.5rem">
      <svg class="svg-icon" width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="#14532d" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
        <rect x="3" y="4" width="18" height="18" rx="2" ry="2"/><line x1="16" y1="2" x2="16" y2="6"/>
        <line x1="8" y1="2" x2="8" y2="6"/><line x1="3" y1="10" x2="21" y2="10"/>
      </svg>
      <span style="font-size:1.3rem; font-weight:600;">Jadwal Terdekat</span>
    </div>
    """, unsafe_allow_html=True)
=======
        except requests.exceptions.ConnectionError:
            st.error("Tidak dapat terhubung ke server.")
        except requests.exceptions.Timeout:
            st.error("Koneksi timeout.")
        except Exception as e:
            st.error(f"Terjadi kesalahan: {str(e)}")
>>>>>>> f585c790da4d1a3b11f6ab84c1283f44dd5e6219

    # Jadwal Terdekat
    st.markdown("---")
    st.markdown('<div style="display:flex;align-items:center;gap:8px;"><i class="fa-solid fa-calendar-week" style="color:#14532d;"></i><span style="font-size:1.2rem;font-weight:700;color:#14532d;">Jadwal Terdekat</span></div>', unsafe_allow_html=True)
    try:
        resp_jadwal=api_get("jadwal-terdekat/")
        if resp_jadwal.status_code==200:
            jadwal=resp_jadwal.json()
            if jadwal:
                df=pd.DataFrame(jadwal)
                df.columns=['Tanggal','Lokasi','Kegiatan','Penyerta']
                st.dataframe(df, use_container_width=True, hide_index=True)
            else: st.info("Belum ada jadwal terdekat.")
        else: st.warning("Gagal memuat jadwal.")
    except: st.warning("Gagal memuat jadwal.")

# ════════════════════════════════════════
# LOGIN ADMIN
# ════════════════════════════════════════
elif st.session_state.page=="admin_login":
    _,col_m,_=st.columns([1,2,1])
    with col_m:
        st.markdown("<h3 style='text-align:center;color:#14532d;'><i class='fa-solid fa-lock'></i> Login Admin</h3>", unsafe_allow_html=True)
        with st.form("login_form"):
            username=st.text_input("Username")
            password=st.text_input("Password", type="password")
            if st.form_submit_button("Masuk"):
                if not username or not password: show_notif("Isi username dan password.","error")
                else:
                    resp=api_post("login/",{"username":username,"password":password})
                    if resp.status_code==200:
                        d=resp.json()
                        st.session_state.logged_in=True; st.session_state.token=d['token']
                        st.session_state.username=d.get('username',username)
                        st.session_state.page="admin_dashboard"
                        st.query_params["token"]=d['token']; st.rerun()
                    elif resp.status_code==429: show_notif("Terlalu banyak percobaan.","error")
                    else: show_notif(resp.json().get('error','Login gagal'),"error")

# ════════════════════════════════════════
# DASHBOARD ADMIN
# ════════════════════════════════════════
elif st.session_state.page=="admin_dashboard":
    if not st.session_state.logged_in or not st.session_state.token:
        st.warning("Login dulu."); st.session_state.page="admin_login"; st.rerun()

    st.markdown(f'<div style="display:flex;align-items:center;gap:10px;margin-bottom:1rem;"><i class="fa-solid fa-gauge" style="color:#14532d;"></i><span style="font-size:1.4rem;font-weight:700;color:#14532d;">Dashboard Admin</span><span style="margin-left:auto;background:#dcfce7;color:#15803d;font-weight:600;padding:4px 12px;border-radius:999px;">{st.session_state.username}</span></div>', unsafe_allow_html=True)

    tab1,tab2,tab3,tab4,tab5,tab6=st.tabs([
        "Input Manual","Google Sheet","Pencarian","Kelola Data","History","Randomize (Dalam Gedung)"
    ])

    # ── Tab 1: Input Manual ──
    with tab1:
        with st.form("input_form"):
<<<<<<< HEAD
            col1, col2 = st.columns(2)
            with col1:
                tgl = st.date_input("Tanggal")
                lokasi = st.text_input("Lokasi")
            with col2:
                kegiatan = st.text_input("Nama Kegiatan")
                penyerta_terpilih = st.multiselect(
                    "Pilih Penyerta (bisa lebih dari satu)",
                    options=DAFTAR_NAMA,
                    placeholder="Ketik nama atau pilih dari daftar"
                )
                if not penyerta_terpilih:
                    penyerta = st.text_area("Atau tulis manual (pisahkan dengan ;)", height=80)
                else:
                    penyerta = "; ".join(penyerta_terpilih)
                    st.caption(f"Penyerta: {penyerta}")
=======
            c1,c2=st.columns(2)
            with c1: tgl=st.date_input("Tanggal"); lokasi=st.text_input("Lokasi")
            with c2: kegiatan=st.text_input("Nama Kegiatan")
            penyerta_terpilih=st.multiselect("Pilih Penyerta", options=DAFTAR_NAMA)
            if not penyerta_terpilih:
                penyerta=st.text_area("Atau tulis manual (pisahkan ;)", height=80)
            else:
                penyerta="; ".join(penyerta_terpilih); st.caption(f"Penyerta: {penyerta}")
>>>>>>> f585c790da4d1a3b11f6ab84c1283f44dd5e6219
            if st.form_submit_button("Simpan"):
                if not lokasi or not kegiatan: show_notif("Lokasi dan Kegiatan wajib diisi.","error")
                else:
<<<<<<< HEAD
                    try:
                        resp = api_post("kegiatan/", {"tanggal": str(tgl), "lokasi": lokasi, "kegiatan": kegiatan, "penyerta": penyerta}, auth=True)
                        if resp.status_code == 201:
                            st.session_state.notif = "Data berhasil disimpan!"
                            st.rerun()
                        else:
                            st.error(f"Gagal menyimpan: {resp.text}")
                    except Exception as e:
                        st.error(f"Error: {e}")
=======
                    resp=api_post("kegiatan/",{"tanggal":str(tgl),"lokasi":lokasi,"kegiatan":kegiatan,"penyerta":penyerta},auth=True)
                    if resp.status_code==201: st.session_state.notif="Data berhasil disimpan!"; st.rerun()
                    else: show_notif(f"Gagal: {resp.text}","error")
>>>>>>> f585c790da4d1a3b11f6ab84c1283f44dd5e6219

    # ── Tab 2: Google Sheet ──
    with tab2:
        st.markdown('<h4><i class="fa-solid fa-cloud-arrow-up"></i> Sync Google Spreadsheet</h4>', unsafe_allow_html=True)
        st.info("Kolom yang dibutuhkan: **tanggal**, **lokasi**, **kegiatan**, **penyerta**  \n"
                "Nama dengan koma (gelar) dan nama jamak dipisah titik koma (;) didukung otomatis.")
        st.warning("Mode **Ganti semua** akan menghapus seluruh data lama sebelum mengisi ulang. "
                   "Gunakan **Tambahkan** jika hanya ingin menambah data baru.")

        def normalize_date(date_str):
            if not date_str:
                return None
            try:
                parsed = pd.to_datetime(date_str, dayfirst=True, errors='coerce')
                if pd.notna(parsed):
                    return parsed.strftime('%Y-%m-%d')
                parsed = pd.to_datetime(date_str, dayfirst=False, errors='coerce')
                if pd.notna(parsed):
                    return parsed.strftime('%Y-%m-%d')
            except Exception:
                pass
            return date_str

        def parse_google_sheet_csv(raw_text):
            import csv as _csv
            lines = raw_text.strip().splitlines()
            if not lines:
                return []
            header_parts = [h.strip().lower() for h in lines[0].split(',')]
            col_map = {}
            for i, h in enumerate(header_parts):
                if 'tanggal' in h or h == 'date':
                    col_map['tanggal'] = i
                elif 'lokasi' in h or h in ('location', 'tempat'):
                    col_map['lokasi'] = i
                elif 'kegiatan' in h or 'activity' in h or 'nama kegiatan' in h:
                    col_map['kegiatan'] = i
                elif 'penyerta' in h or 'pelaksana' in h or 'peserta' in h or 'petugas' in h:
                    col_map['penyerta'] = i
            required = ['tanggal', 'lokasi', 'kegiatan', 'penyerta']
            missing = [k for k in required if k not in col_map]
            if missing:
                raise ValueError(f"Kolom tidak ditemukan: {missing}. Header terdeteksi: {header_parts}")
            peny_idx = col_map['penyerta']
            results = []
            for line in lines[1:]:
                line = line.strip()
                if not line:
                    continue
                try:
                    row = list(_csv.reader([line]))[0]
                except Exception:
                    row = []
                if len(row) >= 4:
                    if len(row) > peny_idx + 1:
                        penyerta_val = ', '.join(r.strip() for r in row[peny_idx:])
                    else:
                        penyerta_val = row[peny_idx].strip()
                    tgl_val = row[col_map['tanggal']].strip()
                    lok_val = row[col_map['lokasi']].strip()
                    keg_val = row[col_map['kegiatan']].strip()
                else:
                    parts = line.split(',', peny_idx)
                    if len(parts) <= peny_idx:
                        continue
                    tgl_val = parts[col_map.get('tanggal', 0)].strip()
                    lok_val = parts[col_map.get('lokasi', 1)].strip()
                    keg_val = parts[col_map.get('kegiatan', 2)].strip()
                    penyerta_val = parts[peny_idx].strip().strip('"')
                tgl_normal = normalize_date(tgl_val)
                if tgl_normal and keg_val:
                    results.append({
                        'tanggal':  tgl_normal,
                        'lokasi':   lok_val,
                        'kegiatan': keg_val,
                        'penyerta': penyerta_val,
                    })
            return results

        def sync_from_url(csv_url, mode, auth=True):
            try:
                resp = requests.get(csv_url, timeout=15)
                resp.raise_for_status()
                raw = resp.text
            except Exception as e:
                raise ValueError(f"Gagal mengunduh CSV: {e}")
            records = parse_google_sheet_csv(raw)
            if not records:
                raise ValueError("Tidak ada data valid di spreadsheet.")
            if mode == 'replace':
                try:
                    r_all = api_get("kegiatan/", auth=True)
                    if r_all.status_code == 200:
                        ids_all = [item['id'] for item in r_all.json()]
                        if ids_all:
                            api_post("kegiatan/bulk-delete/", {"ids": ids_all}, auth=True)
                except Exception:
                    pass
            existing_set = set()
            if mode == 'append':
                try:
                    r_ex = api_get("kegiatan/", auth=True)
                    if r_ex.status_code == 200:
                        for item in r_ex.json():
                            existing_set.add((item['tanggal'], item['lokasi'], item['kegiatan']))
                except Exception:
                    pass
            saved = skipped = errors = 0
            for rec in records:
                key = (rec['tanggal'], rec['lokasi'], rec['kegiatan'])
                if mode == 'append' and key in existing_set:
                    skipped += 1
                    continue
                try:
                    r = api_post("kegiatan/", rec, auth=True)
                    if r.status_code == 201:
                        saved += 1
                        existing_set.add(key)
                    else:
                        errors += 1
                except Exception:
                    errors += 1
            return saved, skipped, errors

        if st.session_state.last_csv_url:
            url_display = st.session_state.last_csv_url
            short_url = url_display if len(url_display) <= 60 else url_display[:57] + "..."
            st.markdown(
                f"""<div style="background:#f3f4f6;border:1px solid #d1d5db;border-radius:8px;
                    padding:0.55rem 0.9rem;margin-bottom:0.75rem;display:flex;
                    align-items:center;gap:8px;word-break:break-all;">
                  <i class="fa-solid fa-link" style="color:#6b7280;flex-shrink:0;"></i>
                  <a href="{url_display}" target="_blank"
                     style="font-size:0.83rem;color:#374151;text-decoration:none;flex:1;">
                     {short_url}
                  </a>
                </div>""",
                unsafe_allow_html=True,
            )
            btn_u, btn_h, btn_s = st.columns([1, 1, 2])
            with btn_u:
                if st.button("Update Link", key="btn_update_link"):
                    st.session_state.show_csv_input = True
            with btn_h:
                if st.button("Hapus Link", key="btn_hapus_link"):
                    st.session_state.last_csv_url = ""
                    st.session_state.show_csv_input = False
                    st.rerun()
<<<<<<< HEAD
            show_input = st.session_state.get('show_csv_input', False)
            if show_input:
                csv_url = st.text_input("Link CSV terpublikasi (baru)", key="csv_url_new")
                sync_mode = st.radio("Mode", ['append', 'replace'], horizontal=True, format_func=lambda x: "Tambahkan saja" if x=='append' else "Ganti semua")
                if st.button("Simpan & Sync"):
                    if csv_url:
                        with st.spinner("Sync..."):
                            resp = api_post("sync-sheets/", {"csv_url": csv_url, "mode": sync_mode}, auth=True)
                            if resp.status_code == 200:
                                st.session_state.last_csv_url = csv_url
                                st.session_state.show_csv_input = False
                                st.session_state.notif = "Data Google Sheet berhasil disinkronkan!"
=======
            with btn_s:
                sync_mode_rerun = st.selectbox(
                    "Mode",
                    ['append', 'replace'],
                    format_func=lambda x: "Tambahkan (aman)" if x == 'append' else "Ganti semua (hapus data lama)",
                    key="sync_mode_rerun",
                )
                if st.button("Sync Sekarang", use_container_width=True, key="btn_sync_rerun"):
                    with st.spinner("Mengambil & menyimpan data dari spreadsheet..."):
                        try:
                            saved, skipped, errors = sync_from_url(st.session_state.last_csv_url, sync_mode_rerun)
                            if saved > 0:
                                st.session_state.notif = (
                                    f"Sync berhasil! {saved} disimpan"
                                    + (f", {skipped} dilewati (duplikat)" if skipped else "")
                                    + (f", {errors} gagal" if errors else "")
                                )
>>>>>>> f585c790da4d1a3b11f6ab84c1283f44dd5e6219
                                st.rerun()
                            else:
                                show_notif(f"Tidak ada data baru. {skipped} duplikat, {errors} error.", "info")
                        except Exception as ex:
                            show_notif(f"Gagal: {ex}", "error")

            if st.session_state.get('show_csv_input'):
                st.markdown("---")
                csv_url_new = st.text_input(
                    "Link CSV baru",
                    key="csv_url_new",
                    help="Buka Google Sheet → File → Share → Publish to web → Format CSV → Salin link",
                )
                sync_mode_new = st.radio(
                    "Mode Sync",
                    ['append', 'replace'],
                    horizontal=True,
                    format_func=lambda x: "Tambahkan (aman)" if x == 'append' else "Ganti semua (hapus data lama)",
                    key="sync_mode_new",
                )
                if st.button("Simpan & Sync", key="btn_save_new_link"):
                    if csv_url_new:
                        with st.spinner("Mengambil & menyimpan data dari spreadsheet..."):
                            try:
                                saved, skipped, errors = sync_from_url(csv_url_new, sync_mode_new)
                                st.session_state.last_csv_url  = csv_url_new
                                st.session_state.show_csv_input = False
                                st.session_state.notif = (
                                    f"Sync berhasil! {saved} disimpan"
                                    + (f", {skipped} dilewati" if skipped else "")
                                    + (f", {errors} gagal" if errors else "")
                                )
                                st.rerun()
                            except Exception as ex:
                                show_notif(f"Gagal: {ex}", "error")
                    else:
                        show_notif("Masukkan link CSV terlebih dahulu.", "error")
        else:
<<<<<<< HEAD
            csv_url = st.text_input("Link CSV terpublikasi", key="csv_url_blank")
            sync_mode = st.radio("Mode", ['append', 'replace'], horizontal=True, format_func=lambda x: "Tambahkan saja" if x=='append' else "Ganti semua")
            if st.button("Ambil & Simpan Data"):
                if csv_url:
                    with st.spinner("Sync..."):
                        resp = api_post("sync-sheets/", {"csv_url": csv_url, "mode": sync_mode}, auth=True)
                        if resp.status_code == 200:
                            st.session_state.last_csv_url = csv_url
                            st.session_state.notif = "Data Google Sheet berhasil disinkronkan!"
=======
            csv_url_blank = st.text_input(
                "Link CSV terpublikasi Google Sheet",
                key="csv_url_blank",
                help="Buka Google Sheet → File → Share → Publish to web → Format CSV → Salin link",
            )
            sync_mode_blank = st.radio(
                "Mode Sync",
                ['append', 'replace'],
                horizontal=True,
                format_func=lambda x: "Tambahkan (aman)" if x == 'append' else "Ganti semua (hapus data lama)",
                key="sync_mode_blank",
            )
            if st.button("Ambil & Simpan", key="btn_ambil_simpan"):
                if csv_url_blank:
                    with st.spinner("Mengambil & menyimpan data dari spreadsheet..."):
                        try:
                            saved, skipped, errors = sync_from_url(csv_url_blank, sync_mode_blank)
                            st.session_state.last_csv_url = csv_url_blank
                            st.session_state.notif = (
                                f"Sync berhasil! {saved} disimpan"
                                + (f", {skipped} dilewati" if skipped else "")
                                + (f", {errors} gagal" if errors else "")
                            )
>>>>>>> f585c790da4d1a3b11f6ab84c1283f44dd5e6219
                            st.rerun()
                        except Exception as ex:
                            show_notif(f"Gagal: {ex}", "error")
                else:
                    show_notif("Masukkan link CSV terlebih dahulu.", "error")

          # ── Tab 3: Pencarian ──
    with tab3:
        try:
            rd=api_get("kegiatan/", auth=True)
            semua_data=rd.json() if rd.status_code==200 else []
        except: semua_data=[]
        lokasi_list=sorted(set(d['lokasi'] for d in semua_data if d.get('lokasi')))
        kegiatan_list=sorted(set(d['kegiatan'] for d in semua_data if d.get('kegiatan')))
        penyerta_set=set()
        for d in semua_data:
            for n in parse_penyerta(d.get('penyerta','')): penyerta_set.add(n)
        penyerta_list=sorted(penyerta_set)

        st.markdown('<h4><i class="fa-solid fa-filter"></i> Filter Pencarian</h4>', unsafe_allow_html=True)
        with st.form("search_form"):
            c1,c2,c3=st.columns(3)
            with c1: tgl_src=st.date_input("Tanggal", value=None, key="src_tgl")
            with c2:
                lok_dd=st.selectbox("Lokasi", ["Semua"]+lokasi_list+["Cari manual..."])
                lok_src=st.text_input("Ketik lokasi") if lok_dd=="Cari manual..." else ("" if lok_dd=="Semua" else lok_dd)
            with c3:
                keg_dd=st.selectbox("Kegiatan", ["Semua"]+kegiatan_list+["Cari manual..."])
                keg_src=st.text_input("Ketik kegiatan") if keg_dd=="Cari manual..." else ("" if keg_dd=="Semua" else keg_dd)
            peny_dd=st.selectbox("Penyerta", ["Semua"]+penyerta_list+["Cari manual..."])
            peny_src=st.text_input("Ketik penyerta") if peny_dd=="Cari manual..." else ("" if peny_dd=="Semua" else peny_dd)
            if st.form_submit_button("Cari"):
                prms={}
                if tgl_src: prms["tanggal"]=str(tgl_src)
                if lok_src: prms["lokasi"]=lok_src
                if keg_src: prms["kegiatan"]=keg_src
                if peny_src: prms["penyerta"]=peny_src
                if not prms: st.warning("Isi minimal satu filter.")
                else:
                    try:
                        resp=api_get("search-admin/", params=prms, auth=True)
                        if resp.status_code==200:
                            hasil=resp.json()
                            if hasil:
                                groups=defaultdict(list)
                                for item in hasil:
                                    groups[(item['tanggal'],item['lokasi'],item['kegiatan'])].append(item['penyerta'])
                                st.success(f"Ditemukan {len(groups)} grup data")
<<<<<<< HEAD
                                for (tanggal, lokasi, kegiatan), list_penyerta in groups.items():
                                    all_names = []
                                    for p_text in list_penyerta:
                                        all_names.extend(parse_penyerta(p_text))
                                    penyerta_gabung = '<br>'.join(all_names)
                                    st.markdown(f"""
                                    <div class="result-box">
                                        <b>{tanggal}</b> - {lokasi}<br>
                                        {kegiatan}<br>
                                        <small>{penyerta_gabung}</small>
                                    </div>
                                    """, unsafe_allow_html=True)
                            else:
                                st.info("Tidak ditemukan")
                        else:
                            st.error("Gagal mengambil data")
                    except Exception as e:
                        st.error(f"Error: {e}")
        
        # ========== CEK KARYAWAN TIDAK TERJADWAL ==========
        st.markdown("---")
        st.markdown('<h4><i class="fa-solid fa-user-plus"></i> Cek Karyawan yang Tidak Terjadwal (Untuk Jadwal Luar Gedung)</h4>', unsafe_allow_html=True)
        st.caption("Lihat karyawan yang tidak memiliki jadwal dalam gedung pada tanggal tertentu, untuk dialokasikan ke jadwal luar gedung via SPS.")
        
        col_cek1, col_cek2, col_cek3 = st.columns([2, 2, 1])
        with col_cek1:
            tgl_cek = st.date_input("📅 Pilih Tanggal", datetime.now().date(), key="tgl_cek_karyawan")
        with col_cek2:
            # Filter role (opsional)
            role_filter = st.multiselect("Filter Role (opsional, kosongkan untuk semua)", 
                                         ["Dokter", "Perawat", "Bidan", "Promkes", "Sanitarian", "Gizi", "Apoteker", "Lab", "Gigi", "Administrasi"],
                                         default=[])
        with col_cek3:
            st.write("")
            if st.button("🔍 Cek Karyawan Tidak Terjadwal", use_container_width=True, type="primary"):
                tgl_str = str(tgl_cek)
                
                # Nama yang dikecualikan (tidak masuk jadwal manapun)
                NAMA_DIECUALIKAN = [
                    "Isep Deni Herdian, S.Kep.,MMRS",
                    "Isep Suhendar,SKM"
                ]
                
                # Fungsi get karyawan by role
                def get_karyawan_by_role_local(role_keywords):
                    role_map = {
                        'dokter': ['dr.', 'drg.'],
                        'perawat_ners': ['Ners', 'S.Kep', 'Amd.Kep', 'A.Md.Kep'],
                        'bidan': ['Bdn.', 'S.Tr.Keb', 'Am.Keb', 'A.Md.Keb'],
                        'promkes': ['Promosi', 'SKM'],
                        'sanitarian': ['Sanitarian', 'S.K.M', 'A.Md.KL'],
                        'gizi': ['S.Gz', 'A.Md.Gz'],
                        'apoteker': ['Apt', 'S.Farm'],
                        'lab': ['A.Md.AK'],
                        'gigi': ['drg.', 'S.Tr.Kes'],
                        'administrasi': ['S.E', 'S.T', 'S.Kep', 'S.ST', 'SKM', 'AMd.RMIK']
                    }
                    keywords = role_map.get(role_keywords, [])
                    hasil = []
                    for nama in DAFTAR_NAMA:
                        for kw in keywords:
                            if kw.lower() in nama.lower():
                                if nama not in NAMA_DIECUALIKAN:
                                    hasil.append(nama)
                                break
                    return list(set(hasil))
                
                # Ambil semua karyawan berdasarkan role yang difilter
                semua_karyawan = []
                if "Dokter" in role_filter or not role_filter:
                    semua_karyawan.extend(get_karyawan_by_role_local('dokter'))
                if "Perawat" in role_filter or not role_filter:
                    semua_karyawan.extend(get_karyawan_by_role_local('perawat_ners'))
                if "Bidan" in role_filter or not role_filter:
                    semua_karyawan.extend(get_karyawan_by_role_local('bidan'))
                if "Promkes" in role_filter or not role_filter:
                    semua_karyawan.extend(get_karyawan_by_role_local('promkes'))
                if "Sanitarian" in role_filter or not role_filter:
                    semua_karyawan.extend(get_karyawan_by_role_local('sanitarian'))
                if "Gizi" in role_filter or not role_filter:
                    semua_karyawan.extend(get_karyawan_by_role_local('gizi'))
                if "Apoteker" in role_filter or not role_filter:
                    semua_karyawan.extend(get_karyawan_by_role_local('apoteker'))
                if "Lab" in role_filter or not role_filter:
                    semua_karyawan.extend(get_karyawan_by_role_local('lab'))
                if "Gigi" in role_filter or not role_filter:
                    semua_karyawan.extend(get_karyawan_by_role_local('gigi'))
                if "Administrasi" in role_filter or not role_filter:
                    semua_karyawan.extend(get_karyawan_by_role_local('administrasi'))
                
                # Hapus duplikat
                semua_karyawan = list(set(semua_karyawan))
                
                if not semua_karyawan:
                    st.warning("Tidak ada karyawan dengan filter role yang dipilih")
                else:
                    # Ambil semua jadwal pada tanggal tersebut
                    try:
                        resp_jadwal = api_get("kegiatan/", auth=True)
                        if resp_jadwal.status_code == 200:
                            data_jadwal = resp_jadwal.json()
                            jadwal_tanggal = [j for j in data_jadwal if j['tanggal'] == tgl_str]
                            
                            # Ambil semua nama yang sudah terjadwal di tanggal itu
                            karyawan_terjadwal = set()
                            for j in jadwal_tanggal:
                                penyerta_list = parse_penyerta(j['penyerta'])
                                karyawan_terjadwal.update(penyerta_list)
                            
                            # Cari karyawan yang tidak terjadwal
                            karyawan_tidak_terjadwal = [k for k in semua_karyawan if k not in karyawan_terjadwal]
                            
                            # Tampilkan hasil
                            if karyawan_tidak_terjadwal:
                                st.warning(f"📊 **Total karyawan yang TIDAK terjadwal pada {tgl_str}: {len(karyawan_tidak_terjadwal)} orang**")
                                st.info("💡 Karyawan ini bisa dialokasikan ke **JADWAL LUAR GEDUNG** (input manual via SPS)")
                                
                                # Tampilkan dalam bentuk DATAFRAME (rapi!)
                                st.subheader("📋 Daftar Karyawan Tidak Terjadwal")
                                
                                # Urutkan nama
                                karyawan_sorted = sorted(karyawan_tidak_terjadwal)
                                
                                # Buat dataframe
                                df_tidak_terjadwal = pd.DataFrame({
                                    "No": range(1, len(karyawan_sorted) + 1),
                                    "Nama Karyawan": karyawan_sorted
                                })
                                st.dataframe(df_tidak_terjadwal, use_container_width=True, hide_index=True)
                                
                                # Tombol download CSV
                                csv_data = df_tidak_terjadwal.to_csv(index=False)
                                st.download_button(
                                    label="📥 Download Daftar (CSV)",
                                    data=csv_data,
                                    file_name=f"karyawan_tidak_terjadwal_{tgl_str}.csv",
                                    mime="text/csv"
                                )
                                
                                # Juga tampilkan dalam bentuk teks polos (untuk copy)
                                with st.expander("📄 Lihat sebagai teks (bisa dicopy)"):
                                    daftar_nama = '\n'.join([f"{i+1}. {nama}" for i, nama in enumerate(karyawan_sorted)])
                                    st.code(daftar_nama, language="text")
                                    
                            else:
                                st.success(f"✅ Semua karyawan ({len(semua_karyawan)} orang) sudah memiliki jadwal dalam gedung pada tanggal {tgl_str}!")
                        else:
                            st.error("Gagal mengambil data jadwal")
                    except Exception as e:
                        st.error(f"Error: {e}")

   
      # ── Tab 4: Kelola Data ──
=======
                                for (tgl,lok,keg),list_p in groups.items():
                                    names=[n for p in list_p for n in parse_penyerta(p)]
                                    st.markdown(f"<div class='data-card'><b>{tgl}</b> — {lok}<br><b>{keg}</b><br><small>{'<br>'.join(names)}</small></div>", unsafe_allow_html=True)
                            else: st.info("Tidak ditemukan")
                        else: show_notif("Gagal mengambil data","error")
                    except Exception as e: st.error(f"Error: {e}")

        st.markdown("---")
        st.markdown('<h4><i class="fa-solid fa-user-plus"></i> Cek Karyawan Tidak Terjadwal</h4>', unsafe_allow_html=True)
        st.caption("Lihat karyawan tanpa jadwal dalam gedung pada tanggal tertentu.")
        cc1,cc2,cc3=st.columns([2,2,1])
        with cc1: tgl_cek=st.date_input("Pilih Tanggal", datetime.now().date(), key="tgl_cek_karyawan")
        with cc2: role_filter=st.multiselect("Filter Role", ["Dokter","Perawat","Bidan","Promkes","Sanitarian","Gizi","Apoteker","Lab","Gigi","Administrasi"], default=[])
        with cc3:
            if st.button("Cek", use_container_width=True, type="primary"):
                tgl_str=str(tgl_cek)
                def get_kary_local(rk):
                    rm={'dokter':['dr.','drg.'],'perawat_ners':['Ners','S.Kep','Amd.Kep','A.Md.Kep'],
                        'bidan':['Bdn.','S.Tr.Keb','Am.Keb','A.Md.Keb'],'promkes':['Promosi','SKM'],
                        'sanitarian':['Sanitarian','S.K.M','A.Md.KL'],'gizi':['S.Gz','A.Md.Gz'],
                        'apoteker':['Apt','S.Farm'],'lab':['A.Md.AK'],'gigi':['drg.','S.Tr.Kes'],
                        'administrasi':['S.E','S.T','S.Kep','S.ST','SKM','AMd.RMIK']}
                    kw=rm.get(rk,[])
                    return list(set([n for n in DAFTAR_NAMA if n not in NAMA_DIECUALIKAN and any(k.lower() in n.lower() for k in kw)]))
                semua_k=[]
                roles=list(ROLE_MAP.keys()) if not role_filter else [r for r in role_filter if r in ROLE_MAP]
                for r in roles: semua_k.extend(get_kary_local(r))
                semua_k=list(set(semua_k))
                if not semua_k: st.warning("Tidak ada karyawan.")
                else:
                    try:
                        rj=api_get("kegiatan/", auth=True)
                        if rj.status_code==200:
                            jt=[j for j in rj.json() if j['tanggal']==tgl_str]
                            terjadwal=set()
                            for j in jt: terjadwal.update(parse_penyerta(j['penyerta']))
                            tidak=[k for k in semua_k if k not in terjadwal]
                            if tidak:
                                st.warning(f"{len(tidak)} karyawan tidak terjadwal.")
                                df=pd.DataFrame({"No":range(1,len(tidak)+1),"Nama":sorted(tidak)})
                                st.dataframe(df, use_container_width=True, hide_index=True)
                                st.download_button("Download CSV", df.to_csv(index=False).encode('utf-8-sig'), f"tidak_terjadwal_{tgl_str}.csv",'text/csv')
                            else: st.success("Semua sudah terjadwal.")
                    except: st.error("Gagal mengambil data.")

    # ── Tab 4: Kelola Data ──
>>>>>>> f585c790da4d1a3b11f6ab84c1283f44dd5e6219
    with tab4:
        st.subheader("Daftar Semua Kegiatan")
        try:
            resp=api_get("kegiatan/", auth=True)
            if resp.status_code==200:
                data=resp.json()
                if data:
<<<<<<< HEAD
                    df_raw = pd.DataFrame(data)
                    df_raw['tanggal'] = pd.to_datetime(df_raw['tanggal']).dt.date

                    # Grouping untuk tampilan
                    df_grouped = (
                        df_raw.groupby(['tanggal', 'lokasi', 'kegiatan'], sort=False)
                        .agg(penyerta=('penyerta', lambda x: ';\n'.join(x.tolist())))
                        .reset_index()
                    )
                    df_grouped = df_grouped.sort_values('tanggal')

                    # ==================== DOWNLOAD EXCEL RAPI ====================
                    st.markdown("---")
                    st.subheader("📥 Export Data")
                    if st.button("📊 Download Excel Template Rapi", type="primary", use_container_width=True):
                        with st.spinner("Membuat Excel..."):
                            output = io.BytesIO()
                            
                            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                df_export = df_grouped.copy()
                                df_export.rename(columns={
                                    'tanggal': 'Tanggal',
                                    'lokasi': 'Lokasi',
                                    'kegiatan': 'Kegiatan',
                                    'penyerta': 'Penyerta'
                                }, inplace=True)
                                
                                df_export.to_excel(writer, sheet_name='JADWAL_LENGKAP', index=False)
                                
                                # Styling
                                workbook = writer.book
                                worksheet = writer.sheets['JADWAL_LENGKAP']
                                
                                # Warna Header Hijau Tua
                                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                                header_fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
                                header_font = Font(color="FFFFFF", bold=True)
                                thin_border = Border(
                                    left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin')
                                )
                                
                                for cell in worksheet[1]:
                                    cell.fill = header_fill
                                    cell.font = header_font
                                    cell.alignment = Alignment(horizontal="center", vertical="center")
                                
                                # Lebar kolom dan border
                                worksheet.column_dimensions['A'].width = 15
                                worksheet.column_dimensions['B'].width = 25
                                worksheet.column_dimensions['C'].width = 35
                                worksheet.column_dimensions['D'].width = 70
                                
                                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                                    for cell in row:
                                        cell.border = thin_border
                                        if cell.row > 1:
                                            cell.alignment = Alignment(vertical="top", wrap_text=True)

                            output.seek(0)
                            st.download_button(
                                label="✅ Klik untuk Download Excel",
                                data=output,
                                file_name=f"JADWAL_PUSKESMAS_SANGKALI_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )

                    # Tampilan Tabel
                    event = st.dataframe(
                        df_grouped[['tanggal', 'lokasi', 'kegiatan', 'penyerta']],
                        use_container_width=True,
                        hide_index=True,
                        on_select="rerun",
                        selection_mode="multi-row",
                        key="kelola_data"
                    )

                    # (Kode hapus terpilih, edit, hapus per bulan tetap sama seperti sebelumnya)
                    # ... silakan tempel kembali kode hapus, edit, dan hapus per bulan dari kode lama kamu

                else:
                    st.info("Belum ada data")
            else:
                st.error("Gagal memuat data")
        except Exception as e:
            st.error(f"Error: {e}")

       # ── Tab 5: History ──
    with tab5:
        st.subheader("📜 History Kegiatan")
        st.caption("Lihat riwayat kegiatan yang sudah lewat dan yang akan datang")
        
        # Tombol filter
        col_filter1, col_filter2, col_filter3 = st.columns([2, 2, 1])
        with col_filter1:
            filter_status = st.selectbox("Filter Status", ["Semua", "Sudah Lewat", "Hari Ini", "Akan Datang"])
        with col_filter2:
            filter_bulan_history = st.selectbox("Filter Bulan", 
                                               ["Semua", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
                                                "Juli", "Agustus", "September", "Oktober", "November", "Desember"])
        with col_filter3:
            st.write("")
            if st.button("🔄 Refresh", use_container_width=True):
                st.rerun()
        
        try:
            resp = api_get("kegiatan/", auth=True)
            if resp.status_code == 200:
                data = resp.json()
                if data:
                    today = date.today()
                    df = pd.DataFrame(data)
                    df['tanggal_date'] = pd.to_datetime(df['tanggal']).dt.date
                    df = df.sort_values('tanggal_date', ascending=False)
                    
                    # Terapkan filter
                    if filter_status == "Sudah Lewat":
                        df = df[df['tanggal_date'] < today]
                    elif filter_status == "Hari Ini":
                        df = df[df['tanggal_date'] == today]
                    elif filter_status == "Akan Datang":
                        df = df[df['tanggal_date'] > today]
                    
                    if filter_bulan_history != "Semua":
                        bulan_num = ["Januari", "Februari", "Maret", "April", "Mei", "Juni",
                                    "Juli", "Agustus", "September", "Oktober", "November", "Desember"].index(filter_bulan_history) + 1
                        df = df[pd.to_datetime(df['tanggal']).dt.month == bulan_num]
                    
                    if df.empty:
                        st.info("Tidak ada data sesuai filter")
                    else:
                        st.write(f"📊 **Total: {len(df)} kegiatan**")
                        
                        # Tampilkan data dalam bentuk tabel dengan checkbox untuk hapus
                        st.markdown("---")
                        
                        # Pilih multiple kegiatan untuk dihapus
                        kegiatan_options = []
                        for _, row in df.iterrows():
                            tgl_date = row['tanggal_date']
                            if tgl_date < today:
                                status_icon = "📅 (Sudah Lewat)"
                            elif tgl_date == today:
                                status_icon = "🔴 (HARI INI)"
                            else:
                                status_icon = "🟢 (Akan Datang)"
                            kegiatan_options.append(f"{row['tanggal']} | {row['lokasi']} | {row['kegiatan']} | {row['id']} | {status_icon}")
                        
                        selected_items = st.multiselect("Pilih kegiatan yang ingin dihapus", kegiatan_options)
                        
                        col_hapus1, col_hapus2, col_hapus3 = st.columns([1, 2, 1])
                        with col_hapus2:
                            if selected_items and st.button("🗑️ Hapus Terpilih", type="secondary", use_container_width=True):
                                ids_to_delete = []
                                for item in selected_items:
                                    # Extract ID dari string (format: ... | {id} | ...)
                                    parts = item.split(" | ")
                                    if len(parts) >= 4:
                                        try:
                                            id_kegiatan = int(parts[3])
                                            ids_to_delete.append(id_kegiatan)
                                        except:
                                            pass
                                
                                if ids_to_delete:
                                    with st.spinner(f"Menghapus {len(ids_to_delete)} kegiatan..."):
                                        resp_del = api_post("kegiatan/bulk-delete/", {"ids": ids_to_delete}, auth=True)
                                        if resp_del.status_code == 200:
                                            st.session_state.notif = f"✅ Berhasil menghapus {len(ids_to_delete)} kegiatan!"
                                            st.rerun()
                                        else:
                                            st.error("Gagal menghapus data")
                                else:
                                    st.warning("Tidak ada ID yang valid")
                        
                        # Tampilkan data dalam card
                        st.markdown("---")
                        st.subheader("📋 Daftar Kegiatan")
                        
                        for _, row in df.iterrows():
                            tgl_date = row['tanggal_date']
                            if tgl_date < today:
                                cls = "past"
                                status_text = "✅ Sudah Lewat"
                            elif tgl_date == today:
                                cls = "today"
                                status_text = "🔴 HARI INI"
                            elif tgl_date == today + timedelta(days=1):
                                cls = "tomorrow"
                                status_text = "⏰ Besok"
                            else:
                                cls = "future"
                                status_text = "📅 Akan Datang"
                            
                            hari = tgl_date.strftime('%A, %d %B %Y')
                            st.markdown(f"""
                            <div class="result-box {cls}" style="position: relative;">
                                <div style="display: flex; justify-content: space-between; align-items: center;">
                                    <div>
                                        <b>{hari}</b> <span style="font-size: 0.8rem;">({status_text})</span><br>
                                        <b>Lokasi:</b> {row['lokasi']}<br>
                                        <b>Kegiatan:</b> {row['kegiatan']}<br>
                                        <small><b>Penyerta:</b> {row['penyerta']}</small>
                                    </div>
                                </div>
                            </div>
                            """, unsafe_allow_html=True)
                else:
                    st.info("Belum ada data kegiatan")
            else:
                st.error("Gagal memuat data")
        except Exception as e:
            st.error(f"Error: {e}")
        
        # Tombol hapus semua history (khusus yang sudah lewat)
        st.markdown("---")
        st.subheader("🗑️ Hapus Massal")
        
        col_massal1, col_massal2, col_massal3 = st.columns([1, 2, 1])
        with col_massal2:
            if st.button("⚠️ Hapus Semua Kegiatan yang Sudah Lewat", type="secondary", use_container_width=True):
                st.warning("⚠️ Yakin ingin menghapus SEMUA kegiatan yang sudah lewat?")
                
                col_confirm1, col_confirm2, col_confirm3 = st.columns([1, 1, 1])
                with col_confirm1:
                    if st.button("✅ Ya, Hapus", use_container_width=True):
                        try:
                            resp = api_get("kegiatan/", auth=True)
                            if resp.status_code == 200:
                                data = resp.json()
                                today = date.today()
                                ids_to_delete = []
                                for item in data:
                                    tgl_kegiatan = datetime.strptime(item['tanggal'], '%Y-%m-%d').date()
                                    if tgl_kegiatan < today:
                                        ids_to_delete.append(item['id'])
                                
                                if ids_to_delete:
                                    resp_del = api_post("kegiatan/bulk-delete/", {"ids": ids_to_delete}, auth=True)
                                    if resp_del.status_code == 200:
                                        st.session_state.notif = f"✅ Berhasil menghapus {len(ids_to_delete)} kegiatan yang sudah lewat!"
                                        st.rerun()
                                    else:
                                        st.error("Gagal menghapus")
                                else:
                                    st.info("Tidak ada kegiatan yang sudah lewat")
                        except Exception as e:
                            st.error(f"Error: {e}")
                with col_confirm2:
                    if st.button("❌ Batal", use_container_width=True):
                        st.rerun()
=======
                    df_raw=pd.DataFrame(data)
                    df_grouped=(
                        df_raw.groupby(['tanggal','lokasi','kegiatan'], sort=False)
                        .agg(penyerta=('penyerta',lambda x:';\n'.join(x.tolist())), ids=('id',list))
                        .reset_index()
                    )
                    df_grouped.columns=['Tanggal','Lokasi','Kegiatan','Penyerta','IDs']

                    if st.session_state.pending_delete_ids:
                        flat_ids=[i for sub in st.session_state.pending_delete_ids for i in sub]
                        if flat_ids:
                            with st.spinner("Menghapus..."):
                                rd=api_post("kegiatan/bulk-delete/",{"ids":flat_ids},auth=True)
                                if rd.status_code==200:
                                    st.session_state.notif=rd.json().get('message')
                                    st.session_state.pending_delete_ids=[]; st.rerun()
                                else: show_notif("Gagal menghapus","error")

                    event=st.dataframe(
                        df_grouped[['Tanggal','Lokasi','Kegiatan','Penyerta']],
                        use_container_width=True, hide_index=True,
                        on_select="rerun", selection_mode="multi-row", key="kelola_data"
                    )
                    selected_rows=event.selection.rows if event.selection else []
                    if selected_rows:
                        sel_ids=[df_grouped.iloc[i]['IDs'] for i in selected_rows]
                        st.markdown(f"**{len(sel_ids)} grup terpilih**")
                        if st.button("Hapus Data Terpilih"):
                            st.session_state.pending_delete_ids=sel_ids; st.rerun()

                    st.markdown("---")
                    st.markdown('<h4><i class="fa-solid fa-pen-to-square"></i> Edit Data</h4>', unsafe_allow_html=True)
                    opsi_edit=["—"]+[f"{r['Tanggal']} | {r['Lokasi']} | {r['Kegiatan']}" for _,r in df_grouped.iterrows()]
                    pilihan=st.selectbox("Pilih kegiatan", opsi_edit)
                    if pilihan!="—":
                        tp,lp,kp=pilihan.split(" | ")
                        matched=df_grouped[(df_grouped['Tanggal']==tp)&(df_grouped['Lokasi']==lp)&(df_grouped['Kegiatan']==kp)]
                        if not matched.empty:
                            eid=matched.iloc[0]['IDs'][0]
                            if st.button("Muat Data"):
                                rd=api_get(f"kegiatan/{eid}/", auth=True)
                                if rd.status_code==200:
                                    st.session_state.edit_data=rd.json(); st.session_state.edit_id=eid; st.rerun()
                    if st.session_state.edit_data and st.session_state.edit_id:
                        with st.form("edit_form"):
                            tgl_ed=st.date_input("Tanggal", value=pd.to_datetime(st.session_state.edit_data['tanggal']))
                            lok_ed=st.text_input("Lokasi", value=st.session_state.edit_data['lokasi'])
                            keg_ed=st.text_input("Nama Kegiatan", value=st.session_state.edit_data['kegiatan'])
                            peny_ed=st.text_area("Penyerta", value=st.session_state.edit_data['penyerta'])
                            if st.form_submit_button("Update"):
                                ru=api_put(f"kegiatan/{st.session_state.edit_id}/",
                                           {"tanggal":str(tgl_ed),"lokasi":lok_ed,"kegiatan":keg_ed,"penyerta":peny_ed})
                                if ru.status_code==200:
                                    st.session_state.notif="Data diperbarui!"; st.session_state.edit_data=None; st.session_state.edit_id=None; st.rerun()
                                else: show_notif(ru.text,"error")

                    st.markdown("---")
                    st.markdown('<h4><i class="fa-solid fa-calendar-xmark"></i> Hapus per Bulan/Tahun</h4>', unsafe_allow_html=True)
                    cm,cy,cb=st.columns([2,2,1])
                    with cm: bulan=st.selectbox("Bulan", range(1,13), index=datetime.now().month-1)
                    with cy: tahun=st.number_input("Tahun", value=datetime.now().year)
                    with cb:
                        if st.button("Hapus"):
                            rd=api_post("delete-by-date/",{"month":bulan,"year":tahun},auth=True)
                            if rd.status_code==200: st.session_state.notif=rd.json()['message']; st.rerun()
                            else: show_notif("Gagal","error")

                    st.markdown("---")
                    st.markdown('<h5>⬇ Download Data Kegiatan</h5>', unsafe_allow_html=True)
                    cd1,cd2=st.columns(2)
                    with cd1: generate_csv(data,"semua_kegiatan.csv")
                    with cd2:
                        pdf_buf=generate_jadwal_pdf(data)
                        st.download_button("⬇ Download PDF", data=pdf_buf, file_name="semua_kegiatan.pdf",
                                           mime="application/pdf", use_container_width=True)
                else: st.info("Belum ada data")
            else: show_notif("Gagal memuat data","error")
        except Exception as e: st.error(f"Error: {e}")

            # ── Tab 5: History ──
    with tab5:
        st.subheader("History Kegiatan")
        st.caption("Riwayat kegiatan yang sudah lewat dan akan datang")
        hf1,hf2,hf3=st.columns([2,2,1])
        with hf1: filter_status=st.selectbox("Status",["Semua","Sudah Lewat","Hari Ini","Akan Datang"],key="filter_status_history")
        with hf2: filter_bulan=st.selectbox("Bulan",["Semua"]+list(calendar.month_name)[1:],key="filter_bulan_history")
        with hf3:
            if st.button("Refresh", key="refresh_history"): st.rerun()

        try:
            resp=api_get("kegiatan/", auth=True)
            if resp.status_code==200:
                all_data=resp.json()
                if all_data:
                    today=date.today()
                    df=pd.DataFrame(all_data)
                    df['tanggal_date']=pd.to_datetime(df['tanggal']).dt.date
                    df=df.sort_values('tanggal_date', ascending=False)
                    if filter_status=="Sudah Lewat": df=df[df['tanggal_date']<today]
                    elif filter_status=="Hari Ini": df=df[df['tanggal_date']==today]
                    elif filter_status=="Akan Datang": df=df[df['tanggal_date']>today]
                    if filter_bulan!="Semua":
                        bnum=list(calendar.month_name).index(filter_bulan)
                        df=df[pd.to_datetime(df['tanggal']).dt.month==bnum]
                    filtered_data=df.to_dict('records')
                    if df.empty:
                        st.info("Tidak ada data sesuai filter.")
                    else:
                        items=""
                        for _,row in df.iterrows():
                            td=row['tanggal_date']
                            if td<today: sts="Sudah Lewat"; badge="status-past"
                            elif td==today: sts="HARI INI"; badge="status-today"
                            elif td==today+timedelta(days=1): sts="Besok"; badge="status-soon"
                            else: sts="Akan Datang"; badge="status-future"
                            tgl_fmt=pd.to_datetime(row['tanggal']).strftime('%A, %d %B %Y')
                            peny_list = parse_penyerta(str(row['penyerta']))
                            peny_disp = ''.join(f'<span>{p}</span>' for p in peny_list) if peny_list else f'<span>{str(row["penyerta"])[:120]}</span>'
                            items+=f"""<div class="receipt-item">
                                <div class="receipt-item-title">{row['kegiatan']} <span class="receipt-status-badge {badge}">{sts}</span></div>
                                <div class="receipt-item-row"><span class="receipt-item-label">Tanggal</span><span class="receipt-item-value">{tgl_fmt}</span></div>
                                <div class="receipt-item-row"><span class="receipt-item-label">Lokasi</span><span class="receipt-item-value">{row['lokasi']}</span></div>
                                <div class="receipt-item-row"><span class="receipt-item-label">Penyerta</span><div class="receipt-item-value"><div class="receipt-penyerta-list">{peny_disp}</div></div></div>
                            </div>"""
                        # Tampilkan receipt TANPA teks filter di bawah judul
                        st.markdown(f"""<div class="receipt-box">
                            <div class="receipt-header"><h3>History Kegiatan</h3></div>
                            <div class="receipt-body">{items}</div>
                            <div class="receipt-footer">
                                <div class="receipt-barcode">|||||||||||||||||||||||</div>
                                {len(df)} kegiatan &bull; Puskesmas Sangkali
                            </div>
                        </div>""", unsafe_allow_html=True)
                        hd1,hd2=st.columns(2)
                        with hd1: generate_csv(filtered_data,"history_kegiatan.csv")
                        with hd2:
                            pdf_buf=generate_jadwal_pdf(filtered_data)
                            st.download_button("⬇ Download PDF", data=pdf_buf, file_name="history_kegiatan.pdf",
                                               mime="application/pdf", use_container_width=True)
                else: st.info("Belum ada data kegiatan.")
            else: show_notif("Gagal memuat data","error")
        except Exception as e: st.error(f"Error: {e}")

        # Hapus Massal
        st.markdown("---")
        st.subheader("Hapus Massal")
        if not st.session_state.confirm_hapus_massal:
            if st.button("Hapus Semua Kegiatan yang Sudah Lewat", key="btn_hapus_massal"):
                st.session_state.confirm_hapus_massal=True; st.rerun()
        else:
            st.warning("Yakin ingin menghapus **semua** kegiatan yang sudah lewat? Tindakan ini tidak dapat dibatalkan.")
            cy2,cn2,_=st.columns([1,1,2])
            with cy2:
                if st.button("Ya, Hapus Sekarang", key="confirm_ya_hapus", type="primary"):
                    with st.spinner("Menghapus..."):
                        try:
                            r=api_get("kegiatan/", auth=True)
                            if r.status_code==200:
                                today=date.today()
                                ids_del=[item['id'] for item in r.json()
                                         if datetime.strptime(item['tanggal'],'%Y-%m-%d').date()<today]
                                if ids_del:
                                    rd=api_post("kegiatan/bulk-delete/",{"ids":ids_del},auth=True)
                                    if rd.status_code==200:
                                        st.session_state.notif=f"Berhasil menghapus {len(ids_del)} kegiatan!"
                                        st.session_state.confirm_hapus_massal=False; st.rerun()
                                    else: st.error(f"Gagal menghapus. {rd.text}")
                                else: st.info("Tidak ada kegiatan yang sudah lewat."); st.session_state.confirm_hapus_massal=False
                            else: show_notif("Gagal mengambil data.","error")
                        except Exception as e: st.error(f"Error: {e}")
            with cn2:
                if st.button("Batal", key="confirm_batal_hapus"):
                    st.session_state.confirm_hapus_massal=False; st.rerun()
>>>>>>> f585c790da4d1a3b11f6ab84c1283f44dd5e6219

        # ── Tab 6: Randomize (Dalam Gedung) ──
    with tab6:
<<<<<<< HEAD
        st.subheader("🎲 Randomize Jadwal Bulanan (Dalam Gedung)")
        st.caption("Generate jadwal otomatis untuk hari kerja Senin-Sabtu")
        st.warning("⚠️ Pastikan data PIKET PERSALINAN sudah diinput via SPS terlebih dahulu. Jika ada PIKET PERSALINAN MALAM pada suatu hari, maka esok harinya tidak akan digenerate jadwal dalam gedung.")
        
        # Nama yang TIDAK boleh dimasukkan ke jadwal manapun
        NAMA_DIECUALIKAN = [
            "Isep Deni Herdian, S.Kep.,MMRS",
            "Isep Suhendar,SKM"
        ]
        
        col_bulan, col_tahun, col_gen = st.columns([2, 2, 1])
        with col_bulan:
            bulan_pilih = st.selectbox("Bulan", range(1, 13), index=datetime.now().month-1,
                                       format_func=lambda x: ["Januari", "Februari", "Maret", "April", "Mei", "Juni",
                                                              "Juli", "Agustus", "September", "Oktober", "November", "Desember"][x-1])
        with col_tahun:
            tahun_pilih = st.number_input("Tahun", value=datetime.now().year, min_value=2020, max_value=2030)
        with col_gen:
            st.write("")
            minggu_pilih = st.selectbox("Minggu ke-", [1, 2, 3, 4, 5], index=0)
            if st.button("🎲 Generate Jadwal", use_container_width=True, type="primary"):
                with st.spinner(f"Menggenerate jadwal untuk minggu ke-{minggu_pilih}..."):
                    import calendar
                    import random
                    from datetime import timedelta
                    
                    # Filter nama yang dikecualikan
                    def filter_nama(nama_list):
                        return [n for n in nama_list if n not in NAMA_DIECUALIKAN]
                    
                    def get_by_role(keywords):
                        semua = [n for n in DAFTAR_NAMA if any(kw.lower() in n.lower() for kw in keywords)]
                        return filter_nama(semua)
                    
                    # Fungsi untuk cek jadwal piket malam dari database
                    def cek_piket_malam(tanggal):
                        """Cek apakah ada jadwal PIKET PERSALINAN MALAM pada tanggal tersebut"""
                        try:
                            resp = api_get("kegiatan/", auth=True)
                            if resp.status_code == 200:
                                data = resp.json()
                                for item in data:
                                    if item['tanggal'] == tanggal and item['kegiatan'] == 'PIKET PERSALINAN MALAM':
                                        return True
                        except:
                            pass
                        return False
                    
                    # Semua karyawan berdasarkan role (2 nama sudah dikecualikan)
                    semua_dokter = get_by_role(['dr.', 'drg.'])
                    semua_perawat = get_by_role(['Ners', 'S.Kep', 'Amd.Kep', 'A.Md.Kep'])
                    semua_bidan = get_by_role(['Bdn.', 'S.Tr.Keb', 'Am.Keb', 'A.Md.Keb'])
                    semua_promkes = get_by_role(['Promosi', 'SKM'])
                    semua_sanitarian = get_by_role(['Sanitarian', 'S.K.M', 'A.Md.KL'])
                    semua_gizi = get_by_role(['S.Gz', 'A.Md.Gz'])
                    
                    # Pool untuk ILP dan Prolanis
                    pool_ilp = list(set(semua_perawat + semua_bidan + semua_promkes + semua_sanitarian + semua_gizi))
                    
                    # Karyawan tetap (sudah difilter)
                    tetap_pendaftaran = filter_nama(["Winda Siti Sarah, AMd.RMIK", "Pupung Juliana", "Salsa Sabila"])
                    tetap_bpgigi = filter_nama(["drg.Rifan Hanggoro.M.M.R.S", "Endah Setiawati,S.Tr.Kes"])
                    tetap_apotek = filter_nama(["Khilman Husna Pratama, S.Farm.,Apt", "Nova Silpiany Perdany, A.Md.Farm"])
                    tetap_lab = filter_nama(["Vita Tyana Virista, A.Md.AK", "Gina Giovany, A.Md.AK"])
                    tetap_ciangir = "Haeriah, A.Md.Kep"
                    tetap_sumelap = "Ujang Effendi, S.Kep.,Ners"
                    tetap_admin = filter_nama(["Rangga Ismardana Gasbela,S.T", "Yogi Aris Diyanto, S.E"])
                    extra_admin = filter_nama(["Liska Permatasari, S.Kep.,Ners", "Alitsa Nuur Fithri, S.ST", "Andina Dea Priatna, SKM"])
                    
                    # Tracking penggunaan karyawan per MINGGU
                    used_this_week = set()
                    usage_count = {nama: 0 for nama in pool_ilp + semua_dokter + semua_perawat + semua_bidan}
                    
                    def random_pick_with_fairness(lst, count=1, exclude=None):
                        """Pilih karyawan dengan prioritas yang jarang dipakai"""
                        exclude = exclude or set()
                        available = [x for x in lst if x not in exclude and x not in used_this_week]
                        if not available:
                            available = [x for x in lst if x not in exclude]
                        if not available or len(available) < count:
                            return []
                        available.sort(key=lambda x: usage_count.get(x, 0))
                        return random.sample(available[:min(count*3, len(available))], min(count, len(available)))
                    
                    # Dapatkan minggu tertentu dari bulan
                    cal = calendar.monthcalendar(tahun_pilih, bulan_pilih)
                    minggu_index = minggu_pilih - 1
                    if minggu_index >= len(cal):
                        st.error(f"Bulan ini hanya memiliki {len(cal)} minggu")
                        st.stop()
                    
                    week = cal[minggu_index]
                    work_days = []
                    for day_idx, day in enumerate(week):
                        if day != 0 and day_idx < 6:
                            work_days.append(day)
                    
                    if not work_days:
                        st.warning(f"Minggu ke-{minggu_pilih} tidak memiliki hari kerja")
                        st.stop()
                    
                    hari_names = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu"]
                    jadwal_baru = []
                    
                    # Reset penggunaan untuk minggu ini
                    used_this_week = set()
                    skipped_days = []
                    
                    for tgl in work_days:
                        tgl_str = f"{tahun_pilih}-{bulan_pilih:02d}-{tgl:02d}"
                        tgl_obj = datetime(tahun_pilih, bulan_pilih, tgl)
                        nama_hari = hari_names[tgl_obj.weekday()]
                        
                        # CEK BENTROK: Jika H-1 ada piket malam, SKIP hari ini
                        tgl_sebelum = (tgl_obj - timedelta(days=1)).strftime('%Y-%m-%d')
                        if cek_piket_malam(tgl_sebelum):
                            skipped_days.append(tgl_str)
                            continue
                        
                        used_today = set()
                        
                        # PENDAFTARAN
                        if tetap_pendaftaran:
                            jadwal_baru.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': 'PENDAFTARAN', 'penyerta': '; '.join(tetap_pendaftaran)})
                            used_today.update(tetap_pendaftaran)
                        
                        # SKRINING ILP 1 & 2
                        for i in range(1, 3):
                            p = random_pick_with_fairness(pool_ilp, 1, used_today)
                            if p:
                                used_today.add(p[0])
                                used_this_week.add(p[0])
                                usage_count[p[0]] = usage_count.get(p[0], 0) + 1
                                jadwal_baru.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': f'SKRINING ILP {i}', 'penyerta': p[0]})
                        
                        # POLI PROLANIS (3 orang)
                        p = random_pick_with_fairness(pool_ilp, 3, used_today)
                        if len(p) >= 3:
                            used_today.update(p)
                            used_this_week.update(p)
                            for nama in p:
                                usage_count[nama] = usage_count.get(nama, 0) + 1
                            jadwal_baru.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': 'POLI PROLANIS', 'penyerta': '; '.join(p)})
                        
                        # KLASTER DEWASA-LANSIA 1 & 2
                        for i in range(1, 3):
                            dok = random_pick_with_fairness(semua_dokter, 1, used_today)
                            if dok:
                                used_today.add(dok[0])
                                used_this_week.add(dok[0])
                                usage_count[dok[0]] = usage_count.get(dok[0], 0) + 1
                                jadwal_baru.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': f'KLASTER DEWASA-LANSIA {i}', 'penyerta': dok[0]})
                            else:
                                per = random_pick_with_fairness(semua_perawat, 1, used_today)
                                if per:
                                    used_today.add(per[0])
                                    used_this_week.add(per[0])
                                    usage_count[per[0]] = usage_count.get(per[0], 0) + 1
                                    jadwal_baru.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': f'KLASTER DEWASA-LANSIA {i}', 'penyerta': per[0]})
                        
                        # KLASTER IBU KIA & USG (2 bidan + 1 dokter)
                        b = random_pick_with_fairness(semua_bidan, 2, used_today)
                        d = random_pick_with_fairness(semua_dokter, 1, used_today)
                        used_today.update(b)
                        used_this_week.update(b)
                        for nama in b:
                            usage_count[nama] = usage_count.get(nama, 0) + 1
                        if d:
                            used_today.add(d[0])
                            used_this_week.add(d[0])
                            usage_count[d[0]] = usage_count.get(d[0], 0) + 1
                        jadwal_baru.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': 'KLASTER IBU KIA & USG', 'penyerta': f"{'; '.join(b)}; {d[0] if d else 'Tidak ada'}"})
                        
                        # KLASTER ANAK (2 bidan + 1 dokter)
                        b = random_pick_with_fairness(semua_bidan, 2, used_today)
                        d = random_pick_with_fairness(semua_dokter, 1, used_today)
                        used_today.update(b)
                        used_this_week.update(b)
                        for nama in b:
                            usage_count[nama] = usage_count.get(nama, 0) + 1
                        if d:
                            used_today.add(d[0])
                            used_this_week.add(d[0])
                            usage_count[d[0]] = usage_count.get(d[0], 0) + 1
                        jadwal_baru.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': 'KLASTER ANAK', 'penyerta': f"{'; '.join(b)}; {d[0] if d else 'Tidak ada'}"})
                        
                        # R. IMUNISASI (Kamis, 2 bidan)
                        if nama_hari == "Kamis":
                            b = random_pick_with_fairness(semua_bidan, 2, used_today)
                            used_today.update(b)
                            used_this_week.update(b)
                            for nama in b:
                                usage_count[nama] = usage_count.get(nama, 0) + 1
                            jadwal_baru.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': 'R. IMUNISASI', 'penyerta': '; '.join(b) if b else 'Tidak ada'})
                        
                        # R. TINDAKAN (1 perawat)
                        p = random_pick_with_fairness(semua_perawat, 1, used_today)
                        if p:
                            used_today.add(p[0])
                            used_this_week.add(p[0])
                            usage_count[p[0]] = usage_count.get(p[0], 0) + 1
                            jadwal_baru.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': 'R. TINDAKAN', 'penyerta': p[0]})
                        
                        # BP GIGI, APOTEK, LAB
                        if tetap_bpgigi:
                            jadwal_baru.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': 'BP GIGI', 'penyerta': '; '.join(tetap_bpgigi)})
                            used_today.update(tetap_bpgigi)
                        if tetap_apotek:
                            jadwal_baru.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': 'APOTEK', 'penyerta': '; '.join(tetap_apotek)})
                            used_today.update(tetap_apotek)
                        if tetap_lab:
                            jadwal_baru.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': 'LAB', 'penyerta': '; '.join(tetap_lab)})
                            used_today.update(tetap_lab)
                        
                        # R. TB (hanya Selasa, hanya Mutia Wulansari)
                        if nama_hari == "Selasa":
                            petugas_tb = "Mutia Wulansari.,S.Kep.,Ners"
                            if petugas_tb not in NAMA_DIECUALIKAN:
                                jadwal_baru.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': 'R. TB', 'penyerta': petugas_tb})
                                used_today.add(petugas_tb)
                                used_this_week.add(petugas_tb)
                                usage_count[petugas_tb] = usage_count.get(petugas_tb, 0) + 1
                        
                        # ADMINISTRASI
                        extra = random_pick_with_fairness(extra_admin, 2)
                        semua_admin = tetap_admin + extra
                        if semua_admin:
                            jadwal_baru.append({'tanggal': tgl_str, 'lokasi': 'Dalam Gedung', 'kegiatan': 'ADMINISTRASI', 'penyerta': '; '.join(semua_admin)})
                            used_today.update(semua_admin)
                            used_this_week.update(semua_admin)
                            for nama in semua_admin:
                                usage_count[nama] = usage_count.get(nama, 0) + 1
                        
                        # PUSTU
                        jadwal_baru.append({'tanggal': tgl_str, 'lokasi': 'Pustu Ciangir', 'kegiatan': 'PELAYANAN PUSTU', 'penyerta': tetap_ciangir})
                        jadwal_baru.append({'tanggal': tgl_str, 'lokasi': 'Pustu Sumelap', 'kegiatan': 'PELAYANAN PUSTU', 'penyerta': tetap_sumelap})
                        used_today.add(tetap_ciangir)
                        used_today.add(tetap_sumelap)
                    
                    # Tampilkan hari yang di-skip karena piket malam
                    if skipped_days:
                        st.info(f"⚠️ Hari berikut ini TIDAK digenerate karena H-1 ada PIKET PERSALINAN MALAM: {', '.join(skipped_days)}")
                    
                    # Simpan ke database
                    saved = 0
                    progress_text = st.empty()
                    progress_bar = st.progress(0)
                    
                    for i, j in enumerate(jadwal_baru):
                        progress_text.text(f"Menyimpan {i+1} dari {len(jadwal_baru)}: {j['kegiatan']}")
                        progress_bar.progress((i+1)/len(jadwal_baru))
                        try:
                            resp_cek = api_get("kegiatan/", auth=True)
                            exists = False
                            if resp_cek.status_code == 200:
                                existing = resp_cek.json()
                                exists = any(e['tanggal'] == j['tanggal'] and e['kegiatan'] == j['kegiatan'] for e in existing)
                            if not exists:
                                resp = api_post("kegiatan/", j, auth=True)
                                if resp.status_code == 201:
                                    saved += 1
                        except Exception as e:
                            pass
                    
                    progress_text.empty()
                    progress_bar.empty()
                    
                    if saved > 0:
                        st.success(f"✅ Berhasil generate {saved} jadwal untuk minggu ke-{minggu_pilih}!")
                        st.balloons()
                        st.rerun()
                    else:
                        st.error("Gagal menyimpan jadwal. Cek koneksi ke server Django.")
        
        st.markdown("---")
        st.subheader("📋 Hasil Generate Jadwal")
        
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            bulan_filter = st.selectbox("Filter Bulan", range(1, 13), index=datetime.now().month-1,
                                       format_func=lambda x: ["Januari","Februari","Maret","April","Mei","Juni",
                                                              "Juli","Agustus","September","Oktober","November","Desember"][x-1],
                                       key="filter_bulan")
        with col_f2:
            tahun_filter = st.number_input("Filter Tahun", value=datetime.now().year, key="filter_tahun")
        
        try:
            resp = api_get("kegiatan/", auth=True)
            if resp.status_code == 200:
                data = resp.json()
                if data:
                    df = pd.DataFrame(data)
                    df['tanggal'] = pd.to_datetime(df['tanggal'])
                    df_filter = df[(df['tanggal'].dt.month == bulan_filter) & (df['tanggal'].dt.year == tahun_filter)]
                    df_filter = df_filter.sort_values('tanggal')
                    if not df_filter.empty:
                        st.dataframe(df_filter[['tanggal', 'lokasi', 'kegiatan', 'penyerta']], use_container_width=True, hide_index=True)
                    else:
                        st.info(f"Belum ada jadwal untuk {bulan_filter}/{tahun_filter}")
                else:
                    st.info("Belum ada data")
        except:
            st.warning("Gagal mengambil data")
# ─── Footer ───
st.markdown("---")
st.markdown("<div style='text-align:center'>Puskesmas Sangkali &copy; 2026</div>", unsafe_allow_html=True)



=======
        st.subheader("Randomize Jadwal Bulanan (Dalam Gedung)")
        st.caption("Generate jadwal otomatis untuk hari kerja Senin-Sabtu")
        st.warning("Pastikan data PIKET PERSALINAN sudah diinput via SPS terlebih dahulu. Jika ada PIKET PERSALINAN MALAM pada suatu hari, maka esok harinya tidak akan digenerate jadwal dalam gedung.")

        NAMA_DIECUALIKAN_LOC=["Isep Deni Herdian, S.Kep.,MMRS","Isep Suhendar,SKM"]

        rb,ry,rg=st.columns([2,2,1])
        with rb: bulan_pilih=st.selectbox("Bulan", range(1,13), index=datetime.now().month-1, format_func=lambda x:calendar.month_name[x])
        with ry: tahun_pilih=st.number_input("Tahun", value=datetime.now().year, min_value=2020, max_value=2030)
        with rg:
            st.write("")
            minggu_pilih=st.selectbox("Minggu ke-",[1,2,3,4,5], index=0)
            if st.button("Generate Jadwal", use_container_width=True, type="primary"):
                with st.spinner(f"Menggenerate jadwal untuk minggu ke-{minggu_pilih}..."):
                    import calendar as cal_mod
                    import random as rnd_mod
                    from datetime import timedelta as td_mod

                    def fn_local(nl): return [n for n in nl if n not in NAMA_DIECUALIKAN_LOC]
                    def gbr(kws):
                        s=[n for n in DAFTAR_NAMA if any(kw.lower() in n.lower() for kw in kws)]
                        return fn_local(s)
                    def cek_piket_malam(tgl_str):
                        try:
                            r=api_get("kegiatan/", auth=True)
                            if r.status_code==200:
                                for item in r.json():
                                    if item['tanggal']==tgl_str and item['kegiatan']=='PIKET PERSALINAN MALAM': return True
                        except: pass
                        return False

                    sdok=gbr(['dr.','drg.']); sper=gbr(['Ners','S.Kep','Amd.Kep','A.Md.Kep'])
                    sbid=gbr(['Bdn.','S.Tr.Keb','Am.Keb','A.Md.Keb']); sprm=gbr(['Promosi','SKM'])
                    ssan=gbr(['Sanitarian','S.K.M','A.Md.KL']); sgiz=gbr(['S.Gz','A.Md.Gz'])
                    pool_ilp=list(set(sper+sbid+sprm+ssan+sgiz))
                    tp=fn_local(PENDAFTARAN_TETAP); tg=fn_local(BP_GIGI_TETAP)
                    ta=fn_local(APOTEK_TETAP); tl=fn_local(LAB_TETAP)
                    tc=PUSTU_CIANGIR; ts=PUSTU_SUMELAP
                    tadm=fn_local(ADMINISTRASI_TETAP); eadm=fn_local(ADMINISTRASI_EXTRA)
                    used_week=set(); uc={n:0 for n in pool_ilp+sdok+sper+sbid}

                    def rpf(lst, count=1, excl=None):
                        excl=excl or set()
                        av=[x for x in lst if x not in excl and x not in used_week]
                        if not av: av=[x for x in lst if x not in excl]
                        if not av or len(av)<count: return []
                        av.sort(key=lambda x:uc.get(x,0))
                        return rnd_mod.sample(av[:min(count*3,len(av))], min(count,len(av)))

                    cal_m=cal_mod.monthcalendar(tahun_pilih,bulan_pilih)
                    mi=minggu_pilih-1
                    if mi>=len(cal_m): st.error(f"Bulan ini hanya {len(cal_m)} minggu"); st.stop()
                    week=cal_m[mi]
                    work_days=[d for idx,d in enumerate(week) if d!=0 and idx<6]
                    if not work_days: st.warning("Tidak ada hari kerja di minggu ini"); st.stop()

                    hari_names=["Senin","Selasa","Rabu","Kamis","Jumat","Sabtu"]
                    jadwal_baru=[]; skipped=[]

                    for tgl in work_days:
                        tgl_str=f"{tahun_pilih}-{bulan_pilih:02d}-{tgl:02d}"
                        tgl_obj=datetime(tahun_pilih,bulan_pilih,tgl)
                        nh=hari_names[tgl_obj.weekday()]
                        tgl_sbl=(tgl_obj-td_mod(days=1)).strftime('%Y-%m-%d')
                        if cek_piket_malam(tgl_sbl): skipped.append(tgl_str); continue
                        used_today=set()
                        if tp:
                            jadwal_baru.append({'tanggal':tgl_str,'lokasi':'Dalam Gedung','kegiatan':'PENDAFTARAN','penyerta':'; '.join(tp)})
                            used_today.update(tp)
                        for i in range(1,3):
                            p=rpf(pool_ilp,1,used_today)
                            if p: used_today.add(p[0]); used_week.add(p[0]); uc[p[0]]=uc.get(p[0],0)+1
                            jadwal_baru.append({'tanggal':tgl_str,'lokasi':'Dalam Gedung','kegiatan':f'SKRINING ILP {i}','penyerta':p[0] if p else 'Tidak ada'})
                        p=rpf(pool_ilp,3,used_today)
                        if len(p)>=3:
                            used_today.update(p); used_week.update(p)
                            for n in p: uc[n]=uc.get(n,0)+1
                            jadwal_baru.append({'tanggal':tgl_str,'lokasi':'Dalam Gedung','kegiatan':'POLI PROLANIS','penyerta':'; '.join(p)})
                        for i in range(1,3):
                            dok=rpf(sdok,1,used_today)
                            if dok: used_today.add(dok[0]); used_week.add(dok[0]); uc[dok[0]]=uc.get(dok[0],0)+1
                            jadwal_baru.append({'tanggal':tgl_str,'lokasi':'Dalam Gedung','kegiatan':f'KLASTER DEWASA-LANSIA {i}','penyerta':dok[0] if dok else 'Tidak ada'})
                        for label in ['KLASTER IBU KIA & USG','KLASTER ANAK']:
                            b=rpf(sbid,2,used_today); d=rpf(sdok,1,used_today)
                            used_today.update(b); used_week.update(b)
                            for n in b: uc[n]=uc.get(n,0)+1
                            if d: used_today.add(d[0]); used_week.add(d[0]); uc[d[0]]=uc.get(d[0],0)+1
                            jadwal_baru.append({'tanggal':tgl_str,'lokasi':'Dalam Gedung','kegiatan':label,'penyerta':f"{'; '.join(b)}; {d[0] if d else 'Tidak ada'}"})
                        if nh=="Kamis":
                            b=rpf(sbid,2,used_today); used_today.update(b); used_week.update(b)
                            for n in b: uc[n]=uc.get(n,0)+1
                            jadwal_baru.append({'tanggal':tgl_str,'lokasi':'Dalam Gedung','kegiatan':'R. IMUNISASI','penyerta':'; '.join(b) if b else 'Tidak ada'})
                        p=rpf(sper,1,used_today)
                        if p: used_today.add(p[0]); used_week.add(p[0]); uc[p[0]]=uc.get(p[0],0)+1
                        jadwal_baru.append({'tanggal':tgl_str,'lokasi':'Dalam Gedung','kegiatan':'R. TINDAKAN','penyerta':p[0] if p else 'Tidak ada'})
                        if tg: jadwal_baru.append({'tanggal':tgl_str,'lokasi':'Dalam Gedung','kegiatan':'BP GIGI','penyerta':'; '.join(tg)}); used_today.update(tg)
                        if ta: jadwal_baru.append({'tanggal':tgl_str,'lokasi':'Dalam Gedung','kegiatan':'APOTEK','penyerta':'; '.join(ta)}); used_today.update(ta)
                        if tl: jadwal_baru.append({'tanggal':tgl_str,'lokasi':'Dalam Gedung','kegiatan':'LAB','penyerta':'; '.join(tl)}); used_today.update(tl)
                        if nh=="Selasa":
                            ptb="Mutia Wulansari.,S.Kep.,Ners"
                            if ptb not in NAMA_DIECUALIKAN_LOC:
                                jadwal_baru.append({'tanggal':tgl_str,'lokasi':'Dalam Gedung','kegiatan':'R. TB','penyerta':ptb})
                                used_today.add(ptb); used_week.add(ptb); uc[ptb]=uc.get(ptb,0)+1
                        extra=rpf(eadm,2)
                        sadm=tadm+extra
                        if sadm:
                            jadwal_baru.append({'tanggal':tgl_str,'lokasi':'Dalam Gedung','kegiatan':'ADMINISTRASI','penyerta':'; '.join(sadm)})
                            used_today.update(sadm); used_week.update(sadm)
                            for n in sadm: uc[n]=uc.get(n,0)+1
                        jadwal_baru.append({'tanggal':tgl_str,'lokasi':'Pustu Ciangir','kegiatan':'PELAYANAN PUSTU','penyerta':tc})
                        jadwal_baru.append({'tanggal':tgl_str,'lokasi':'Pustu Sumelap','kegiatan':'PELAYANAN PUSTU','penyerta':ts})
                        used_today.add(tc); used_today.add(ts)

                    if skipped: st.info(f"Tidak digenerate (H-1 ada PIKET MALAM): {', '.join(skipped)}")
                    saved=0; ptxt=st.empty(); pbar=st.progress(0)
                    existing_all=[]
                    try:
                        re2=api_get("kegiatan/", auth=True)
                        if re2.status_code==200: existing_all=re2.json()
                    except: pass
                    exist_set={(e['tanggal'],e['kegiatan']) for e in existing_all}
                    for i,j in enumerate(jadwal_baru):
                        ptxt.text(f"Menyimpan {i+1}/{len(jadwal_baru)}: {j['kegiatan']}")
                        pbar.progress((i+1)/len(jadwal_baru))
                        if (j['tanggal'],j['kegiatan']) not in exist_set:
                            try:
                                r=api_post("kegiatan/",j,auth=True)
                                if r.status_code==201: saved+=1; exist_set.add((j['tanggal'],j['kegiatan']))
                            except: pass
                    ptxt.empty(); pbar.empty()
                    if saved>0: st.session_state.notif=f"Berhasil generate {saved} jadwal minggu ke-{minggu_pilih}!"; st.balloons(); st.rerun()
                    else: st.error("Gagal menyimpan jadwal. Cek koneksi ke server Django.")

        st.markdown("---")
        st.subheader("Hasil Generate Jadwal")
        gf1,gf2=st.columns(2)
        with gf1: bulan_filter=st.selectbox("Filter Bulan", range(1,13), index=datetime.now().month-1, format_func=lambda x:calendar.month_name[x], key="filter_bulan")
        with gf2: tahun_filter=st.number_input("Filter Tahun", value=datetime.now().year, key="filter_tahun")
        try:
            resp=api_get("kegiatan/", auth=True)
            if resp.status_code==200:
                data=resp.json()
                if data:
                    df=pd.DataFrame(data)
                    df['tanggal']=pd.to_datetime(df['tanggal'])
                    df_f=df[(df['tanggal'].dt.month==bulan_filter)&(df['tanggal'].dt.year==tahun_filter)].sort_values('tanggal')
                    if not df_f.empty: st.dataframe(df_f[['tanggal','lokasi','kegiatan','penyerta']], use_container_width=True, hide_index=True)
                    else: st.info(f"Belum ada jadwal untuk {bulan_filter}/{tahun_filter}")
                else: st.info("Belum ada data")
        except: st.warning("Gagal mengambil data")

# ── FOOTER ──
st.markdown("""
<div class="puskesmas-footer">
    <div class="footer-grid">
        <div class="footer-section">
            <h4>Puskesmas Sangkali</h4>
            <p><i class="fa-solid fa-location-dot"></i> Jl. Sangkali No. 1, Tamansari<br>Tasikmalaya, Jawa Barat</p>
            <p><i class="fa-solid fa-phone"></i> (0265) 123456</p>
            <p><i class="fa-solid fa-envelope"></i> puskesmas@sangkali.id</p>
        </div>
        <div class="footer-section">
            <h4>Jam Operasional</h4>
            <p><i class="fa-solid fa-clock"></i> Senin–Kamis: 07:30–14:00</p>
            <p><i class="fa-solid fa-clock"></i> Jumat: 07:30–11:30</p>
            <p><i class="fa-solid fa-clock"></i> Sabtu: 07:30–12:00</p>
        </div>
        <div class="footer-section"></div>
    </div>
    <hr class="footer-divider">
    <div class="footer-bottom">&copy; 2026 UPTD Puskesmas Sangkali</div>
</div>
""", unsafe_allow_html=True)
>>>>>>> f585c790da4d1a3b11f6ab84c1283f44dd5e6219
